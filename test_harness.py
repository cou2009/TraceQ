#!/usr/bin/env python3
"""
TraceQ Test Harness — runs ALL samples against ALL BOQs.
Must be run after EVERY engine change. No exceptions.

Usage: python3 test_harness.py
"""
import sys, os, json
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from traceq_engine import TraceQEngine
import openpyxl

BASE = "/sessions/sweet-exciting-heisenberg/mnt/Claude Docs"

SAMPLES = {
    "S1": {
        "dxf_dir": f"{BASE}/Sample 1 - HVAC/Sample HVAC -1/DXF/",
        "boq": f"{BASE}/Sample 1 - HVAC/Sample HVAC -1/HVAC- Bill of quantity.xlsx",
        "boq_sheet": "SEC L- BOQ",
        "boq_parser": "s1",
    },
    "S2": {
        "dxf_files": [f"{BASE}/Sample 2 - HVAC /Sample HVAC -2/HVAC LAYOUTS (1).dxf"],
        "boq": f"{BASE}/Sample 2 - HVAC /Sample HVAC -2/HVAC - Bill of Quantities.xlsx",
        "boq_sheet": "Sec 15 HVAC",
        "boq_parser": "s2s3",
    },
    "S3": {
        "dxf_dir": f"{BASE}/Sample 3 - HVAC /Sample HVAC -3/DXF/",
        "boq": f"{BASE}/Sample 3 - HVAC /Sample HVAC -3/HVAC - Bill of Quantities.xlsx",
        "boq_sheet": "Sec 8 HVAC",
        "boq_parser": "s2s3",
    },
    "S4": {
        "dxf_files": [f"{BASE}/Sample 4 - HVAC /Sample HVAC -4/HVAC.dxf"],
        "boq": f"{BASE}/Sample 4 - HVAC /Sample HVAC -4/HVAC BOQ.xlsx",
        "boq_sheet": "MECHANICAL BOQ",
        "boq_parser": "s4",
    },
    "S5": {
        "dxf_files": [f"{BASE}/Sample 5 - HVAC /Sample 5 - Layout drawings.dxf"],
        "boq": f"{BASE}/Sample 5 - HVAC /Sample 5- HVAC BOQ.xlsx",
        "boq_sheet": "MECHANICAL BOQ - 1",
        "boq_parser": "s5s6",
    },
    "S6": {
        "dxf_files": [f"{BASE}/Sample 6 - HVAC /Sample 6 - Layout drawings.dxf"],
        "boq": f"{BASE}/Sample 6 - HVAC /Sample 6- HVAC BOQ.xlsx",
        "boq_sheet": "MECHANICAL BOQ - 2",
        "boq_parser": "s5s6",
    },
}

# ── BOQ PARSERS ──────────────────────────────────────────────────────────────
# Each sample's BOQ has a different structure. We parse them into a common
# format: list of {"item": str, "qty": float, "unit": str}

def parse_boq_s1(ws):
    """S1: cols B=description, C=unit, D=qty"""
    items = []
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=False):
        vals = {c.column_letter: c.value for c in row if not isinstance(c, openpyxl.cell.cell.MergedCell)}
        desc = str(vals.get('B', '') or '').strip()
        unit = str(vals.get('C', '') or '').strip()
        qty = vals.get('D', None)
        if desc and qty and isinstance(qty, (int, float)) and qty > 0:
            items.append({"item": desc, "qty": float(qty), "unit": unit})
    return items

def parse_boq_s2s3(ws):
    """S2/S3: cols C=description, E=qty, F=unit"""
    items = []
    for row in ws.iter_rows(min_row=10, max_row=ws.max_row, values_only=False):
        vals = {c.column_letter: c.value for c in row if not isinstance(c, openpyxl.cell.cell.MergedCell)}
        desc = str(vals.get('C', '') or '').strip()
        qty = vals.get('E', None)
        unit = str(vals.get('F', '') or '').strip()
        if desc and qty and isinstance(qty, (int, float)) and qty > 0:
            items.append({"item": desc, "qty": float(qty), "unit": unit})
    return items

def parse_boq_s4(ws):
    """S4: cols C=description, D=unit, E=qty"""
    items = []
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=False):
        vals = {c.column_letter: c.value for c in row if not isinstance(c, openpyxl.cell.cell.MergedCell)}
        desc = str(vals.get('C', '') or '').strip()
        unit = str(vals.get('D', '') or '').strip()
        qty = vals.get('E', None)
        if desc and qty and isinstance(qty, (int, float)) and qty > 0:
            items.append({"item": desc, "qty": float(qty), "unit": unit})
    return items

def parse_boq_s5s6(ws):
    """S5/S6: cols C=description, D=unit, E=qty"""
    items = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=False):
        vals = {c.column_letter: c.value for c in row if not isinstance(c, openpyxl.cell.cell.MergedCell)}
        desc = str(vals.get('C', '') or '').strip()
        unit = str(vals.get('D', '') or '').strip()
        qty = vals.get('E', None)
        if desc and qty and isinstance(qty, (int, float)) and qty > 0:
            items.append({"item": desc, "qty": float(qty), "unit": unit})
    return items

BOQ_PARSERS = {
    "s1": parse_boq_s1,
    "s2s3": parse_boq_s2s3,
    "s4": parse_boq_s4,
    "s5s6": parse_boq_s5s6,
}

# ── EQUIPMENT CATEGORY MAPPER ────────────────────────────────────────────────
# Maps BOQ description keywords to our engine's equipment categories

CATEGORY_MAP = {
    # NOTE: Order matters — first match wins. Put more specific patterns first.
    "fcu": ["fcu", "fan coil", "fan-coil", "ducted split", "decorative split"],
    "indoor_unit": ["indoor unit", "vrv-idu", "dx-idu", "vrv-ahu"],
    "outdoor_unit": ["outdoor unit", "vrv-odu"],
    "vrf": ["vrf unit", "vrf outdoor", "vrv outdoor", "vrf", "vrv"],
    "volume_control_damper": ["vcd", "volume control damper", "volume damper", "Ø volume damper"],
    "fire_damper": ["fire damper"],
    "supply_diffuser": ["supply air diffuser", "supply diffuser", "sad ", "sald", "square diffuser"],
    "return_diffuser": ["return air diffuser", "return diffuser", "rad ", "rald", "return air grille", "return grill"],
    "thermostat": ["thermostat", "madoca", " with ts"],
    "flow_bar": ["flow bar", "slot diffuser", "linear slot", "leniar slot", "linear diffuser", "leniar diffuser"],
    "plenum_box": ["plenum box"],
    "grille": ["grille", "slbg", "supply air grille"],
    "air_curtain": ["air curtain"],
    "extract_diffuser": ["extract diffuser", "exhaust diffuser", "exhaust air diffuser", "toilet exhaust", "ead", "disk valve", "disc valve", "ed 150"],
    "flexible_duct": ["flexible duct", "flexible connector"],
    "supply_duct": ["supply & return air gi", "supply air duct", "supply duct", "a.c supply & return"],
    "return_duct": ["return air duct", "return duct"],
    "sound_attenuator": ["sound attenuator", "acoustic elbow"],
    "motorized_damper": ["motorize damper", "motorized damper"],
    "non_return_damper": ["non-return damper", "non return damper", "nrd"],
    "exhaust_fan": ["extract fan", "exhaust fan", "eaf-", "faf-", "kitchen hood extract", "make up air fan", "smoke extract fan", "toilet extract fan", "stairwell pressuri", "smoke extract"],
    "circular_diffuser": ["circular diffuser", "round diffuser"],
    "insulation": ["insulation", "acoustic lin"],
    "louver": ["louver", "sand trap"],
    "fahu": ["fahu", "fresh air handling"],
    "copper_piping": ["copper pip", "refrigerant copper"],
    "air_handling_unit": ["air handling unit", "ahu"],
    "access_door": ["access door", "access panel", "inspection door"],
}

def categorise_boq_item(desc):
    """Map a BOQ description to an equipment category."""
    desc_lower = desc.lower()
    for category, keywords in CATEGORY_MAP.items():
        for kw in keywords:
            if kw in desc_lower:
                return category
    return None

def aggregate_boq(boq_items):
    """Aggregate BOQ items by equipment category, summing quantities."""
    agg = {}
    uncategorised = []
    for item in boq_items:
        cat = categorise_boq_item(item['item'])
        if cat:
            unit = item['unit'].lower().replace('.', '').strip()
            is_countable = unit in ('nos', 'no', 'nos.', 'no.', 'set', 'sets', '')
            if cat not in agg:
                agg[cat] = {"qty": 0, "unit": item['unit'], "countable": is_countable, "items": []}
            agg[cat]["qty"] += item['qty']
            agg[cat]["items"].append(item['item'])
        else:
            uncategorised.append(item)
    return agg, uncategorised

# ── MULTI-VIEW FLOOR DEDUPLICATION ────────────────────────────────────────────
# MEP projects often have paired drawings for the same floor: AC (air conditioning)
# + VE (ventilation) views. Equipment like FCUs appears in both → double counting.
# Detection: match filenames by floor identifier. Dedup: take MAX per equipment
# type per floor, then SUM across floors. Universal, not sample-specific.

import re as _re

def detect_floor_groups(dxf_files):
    """Group DXF files by physical floor. Files covering the same floor from
    different views (AC/VE/VENTILATION) are grouped together.

    Returns: list of groups, where each group is a list of file paths.
    Single files are their own group. Paired files share a group.

    Supports patterns:
      - "AC-100-..." / "VE-100-..." (view prefix + floor number)
      - "...AC LAYOUT GROUND FLOOR..." / "...VENTILATION LAYOUT GROUND FLOOR..."
    """
    assigned = set()
    groups_dict = {}  # floor_key → list of files

    for fpath in dxf_files:
        if fpath in assigned:
            continue
        basename = os.path.basename(fpath).upper()

        # Pattern 1: AC-NNN or VE-NNN prefix (e.g., AC-100-BASEMENT FLOOR PLAN)
        m = _re.match(r'^(AC|VE)-(\d+)', basename)
        if m:
            floor_key = f"P1_{m.group(2)}"
            groups_dict.setdefault(floor_key, []).append(fpath)
            assigned.add(fpath)
            continue

        # Pattern 2: "AC LAYOUT <floor>" or "VENTILATION LAYOUT <floor>"
        m = _re.search(r'(?:^|[\s-])(AC|VENTILATION|VE)\s+LAYOUT\s+(.+?)\.DXF', basename)
        if m:
            floor_desc = m.group(2).strip()
            floor_key = f"P2_{floor_desc}"
            groups_dict.setdefault(floor_key, []).append(fpath)
            assigned.add(fpath)
            continue

        # No pattern match — standalone file
        groups_dict[f"solo_{fpath}"] = [fpath]
        assigned.add(fpath)

    return list(groups_dict.values())


def aggregate_with_dedup(floor_groups, per_file_results):
    """Aggregate equipment counts with multi-view deduplication.

    For floor groups with multiple files: take MAX per equipment type
    (same physical equipment seen from different views).
    For single files: take count directly.
    Sum across all floor groups.

    Args:
        floor_groups: list of [filepath, ...] groups from detect_floor_groups
        per_file_results: dict {filepath: {equip_type: {count, source, ...}}}

    Returns: combined dict {equip_type: {count, source, confidence, ...}}
    """
    combined = {}

    for group in floor_groups:
        if len(group) == 1:
            # Single file — add counts directly
            for equip_type, edata in per_file_results.get(group[0], {}).items():
                count = edata.get('count', 0)
                if equip_type not in combined:
                    combined[equip_type] = {
                        'count': count,
                        'source': edata.get('source', '?'),
                        'confidence': edata.get('confidence', 0),
                        'needs_review': edata.get('needs_review', False),
                        'alternate_counts': edata.get('alternate_counts', {}),
                    }
                else:
                    combined[equip_type]['count'] += count
        else:
            # Multi-view floor — take MAX per equipment type across views
            floor_max = {}
            floor_meta = {}  # Keep metadata from the view with the highest count
            for fpath in group:
                for equip_type, edata in per_file_results.get(fpath, {}).items():
                    count = edata.get('count', 0)
                    if equip_type not in floor_max or count > floor_max[equip_type]:
                        floor_max[equip_type] = count
                        floor_meta[equip_type] = edata

            # Add floor-level MAX to combined totals
            for equip_type, count in floor_max.items():
                if equip_type not in combined:
                    edata = floor_meta[equip_type]
                    combined[equip_type] = {
                        'count': count,
                        'source': edata.get('source', '?'),
                        'confidence': edata.get('confidence', 0),
                        'needs_review': edata.get('needs_review', False),
                        'alternate_counts': edata.get('alternate_counts', {}),
                    }
                else:
                    combined[equip_type]['count'] += count

    return combined


# ── ENGINE RUNNER ─────────────────────────────────────────────────────────────

def run_engine(sample_config):
    """Run engine on all DXF files for a sample, return merged equipment counts.
    Uses multi-view deduplication when floor pairs are detected."""
    engine = TraceQEngine()

    # Get DXF file list
    if 'dxf_dir' in sample_config:
        dxf_dir = sample_config['dxf_dir']
        dxf_files = sorted([
            os.path.join(dxf_dir, f) for f in os.listdir(dxf_dir) if f.lower().endswith('.dxf')
        ])
    else:
        dxf_files = sample_config['dxf_files']

    # Quick scan first file
    scan = engine.quick_scan(dxf_files[0])

    # Full analysis on each file — store per-file results
    per_file_results = {}
    for fpath in dxf_files:
        try:
            result = engine.analyze(fpath)
            rd = result.to_dict()
            file_merged = rd.get('equipment_inventory', rd.get('merged', {}))
            per_file_results[fpath] = file_merged
        except Exception as e:
            print(f"  ⚠️ FAILED on {os.path.basename(fpath)}: {e}")
            per_file_results[fpath] = {}

    # Detect floor pairs and aggregate with deduplication
    floor_groups = detect_floor_groups(dxf_files)
    combined = aggregate_with_dedup(floor_groups, per_file_results)

    return combined, scan, len(dxf_files)

# ── CATEGORY EQUIVALENCES ─────────────────────────────────────────────────────
# Maps BOQ categories to engine categories that should also be checked.
# These represent genuine industry terminology overlaps — the same physical
# equipment gets different category names depending on the consultant/contractor.
# Universal, not sample-specific. Each entry means: "if the BOQ says X, also
# look for Y in the engine results and sum the counts."

CATEGORY_EQUIVALENCES = {
    "fcu": ["indoor_unit"],           # Ducted split units are indoor units functioning as FCUs
    "indoor_unit": ["fcu"],           # Reverse: some BOQs list indoor units, engine may call them FCU
    "vrf": ["outdoor_unit"],          # VRF outdoor units sometimes categorised separately
    "outdoor_unit": ["vrf"],          # Reverse
    "extract_diffuser": ["exhaust_fan"],  # Toilet extract diffusers sometimes classified as exhaust
    "grille": ["return_diffuser", "supply_diffuser"],  # Some BOQs call diffusers "grilles"
}

def get_engine_count(engine_results, boq_cat):
    """Get engine count for a BOQ category, using equivalences as FALLBACK only.
    Logic: use primary count if > 0. If primary is 0, try equivalent categories
    and take the BEST (highest count) match — never sum multiple equivalents.
    Returns (count, source_description)."""
    primary = engine_results.get(boq_cat, {}).get('count', 0)

    if primary > 0:
        return primary, ""

    # Primary is 0 — try equivalents as fallback, pick the best one
    best_count = 0
    best_cat = None
    for equiv_cat in CATEGORY_EQUIVALENCES.get(boq_cat, []):
        equiv_count = engine_results.get(equiv_cat, {}).get('count', 0)
        if equiv_count > best_count:
            best_count = equiv_count
            best_cat = equiv_cat

    if best_count > 0:
        return best_count, f"via {best_cat}={best_count}"

    return 0, ""

def get_matched_engine_categories(boq_agg):
    """Return set of engine categories that are accounted for by BOQ matches
    (including via equivalences). Used to avoid false positive flagging."""
    matched = set()
    for cat in boq_agg:
        matched.add(cat)
        for equiv_cat in CATEGORY_EQUIVALENCES.get(cat, []):
            matched.add(equiv_cat)
    return matched

# ── COMPARISON ────────────────────────────────────────────────────────────────

def compare(engine_results, boq_agg):
    """Compare engine results vs BOQ, return match stats."""
    matches = 0
    mismatches = 0
    not_detected = 0
    false_positives = 0
    verify = 0
    details = []

    # Build set of engine categories covered by BOQ (direct + equivalences)
    covered_categories = get_matched_engine_categories(boq_agg)

    # Check each BOQ category against engine
    for cat, boq_data in sorted(boq_agg.items()):
        engine_count, equiv_source = get_engine_count(engine_results, cat)
        boq_qty = boq_data['qty']
        equiv_note = f"  [{equiv_source}]" if equiv_source else ""

        if not boq_data['countable']:
            details.append(f"  {'⚠️ VERIFY':10s} {cat:30s} BOQ={boq_qty:>8.0f} {boq_data['unit']:8s} Engine={engine_count:>6,}  (unit mismatch){equiv_note}")
            verify += 1
            continue

        if engine_count == 0:
            details.append(f"  {'❌ MISS':10s} {cat:30s} BOQ={boq_qty:>8.0f}  Engine=     0  NOT DETECTED")
            not_detected += 1
        elif abs(engine_count - boq_qty) / max(boq_qty, 1) <= 0.05:
            details.append(f"  {'✅ MATCH':10s} {cat:30s} BOQ={boq_qty:>8.0f}  Engine={engine_count:>6,}{equiv_note}")
            matches += 1
        elif abs(engine_count - boq_qty) / max(boq_qty, 1) <= 0.15:
            details.append(f"  {'⚠️ CLOSE':10s} {cat:30s} BOQ={boq_qty:>8.0f}  Engine={engine_count:>6,}  ({engine_count/boq_qty:.0%}){equiv_note}")
            matches += 0.5  # Half credit
            mismatches += 0.5
        else:
            direction = "OVER" if engine_count > boq_qty else "UNDER"
            details.append(f"  {'❌ ' + direction:10s} {cat:30s} BOQ={boq_qty:>8.0f}  Engine={engine_count:>6,}  ({engine_count/boq_qty:.0%}){equiv_note}")
            mismatches += 1

    # Check for false positives (engine found but not in BOQ, with significant count)
    # Skip categories that are covered by BOQ via equivalences
    for equip, edata in sorted(engine_results.items(), key=lambda x: -x[1]['count']):
        if equip not in covered_categories and edata['count'] > 5:
            details.append(f"  {'🔍 EXTRA':10s} {equip:30s}                Engine={edata['count']:>6,}  (not in BOQ)")
            false_positives += 1

    total_boq_countable = sum(1 for v in boq_agg.values() if v['countable'])
    score = matches / max(total_boq_countable, 1) * 100

    return {
        'matches': matches,
        'mismatches': mismatches,
        'not_detected': not_detected,
        'false_positives': false_positives,
        'verify': verify,
        'total_boq_items': total_boq_countable,
        'score': score,
        'details': details,
    }

# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 90)
    print("  TRACEQ TEST HARNESS — ALL SAMPLES VS ALL BOQs")
    print(f"  Run at: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 90)

    overall_scores = {}

    for sample_name, config in SAMPLES.items():
        print(f"\n{'─' * 90}")
        print(f"  {sample_name}")
        print(f"{'─' * 90}")

        # Run engine
        engine_results, scan, num_files = run_engine(config)
        total_items = sum(e['count'] for e in engine_results.values())
        print(f"  Files: {num_files} | Quick Scan: {scan.overall_score:.1f}% {scan.verdict}")
        print(f"  Engine found: {total_items:,} items across {len(engine_results)} categories")

        # Parse BOQ
        wb = openpyxl.load_workbook(config['boq'], data_only=True)
        ws = wb[config['boq_sheet']]
        parser = BOQ_PARSERS[config['boq_parser']]
        boq_items = parser(ws)
        boq_agg, uncategorised = aggregate_boq(boq_items)
        print(f"  BOQ: {len(boq_items)} line items → {len(boq_agg)} equipment categories")
        if uncategorised:
            print(f"  BOQ uncategorised: {len(uncategorised)} items")
            for u in uncategorised[:5]:
                print(f"    - {u['item'][:60]} ({u['qty']} {u['unit']})")
            if len(uncategorised) > 5:
                print(f"    ... and {len(uncategorised) - 5} more")

        # Compare
        result = compare(engine_results, boq_agg)
        print(f"\n  SCORE: {result['score']:.0f}% ({result['matches']:.0f}/{result['total_boq_items']} countable items)")
        print(f"  Matches={result['matches']:.0f} | Mismatches={result['mismatches']:.0f} | Not Detected={result['not_detected']} | False Positives={result['false_positives']} | Verify={result['verify']}")
        print()
        for line in result['details']:
            print(line)

        overall_scores[sample_name] = {
            'score': result['score'],
            'scan': scan.overall_score,
            'verdict': scan.verdict,
            'matches': result['matches'],
            'total': result['total_boq_items'],
            'false_positives': result['false_positives'],
        }

    # Summary table
    print(f"\n{'=' * 90}")
    print("  OVERALL SCORECARD")
    print(f"{'=' * 90}")
    print(f"  {'Sample':<8} {'Quick Scan':<14} {'Match Score':<16} {'False Pos':<12} {'Verdict'}")
    print(f"  {'─'*8} {'─'*14} {'─'*16} {'─'*12} {'─'*20}")

    total_match = 0
    total_items = 0
    for name, data in overall_scores.items():
        total_match += data['matches']
        total_items += data['total']
        fp_str = f"{data['false_positives']} FPs" if data['false_positives'] else "clean"
        print(f"  {name:<8} {data['scan']:>5.1f}% {data['verdict']:<7} {data['score']:>5.1f}% ({data['matches']:.0f}/{data['total']})   {fp_str:<12} {'✅' if data['score'] >= 50 else '⚠️' if data['score'] >= 25 else '❌'}")

    avg_score = total_match / max(total_items, 1) * 100
    print(f"\n  OVERALL: {avg_score:.1f}% ({total_match:.0f}/{total_items} countable items across all samples)")
    print(f"{'=' * 90}")

    return overall_scores

if __name__ == '__main__':
    main()
