"""
TraceQ â BOQ Risk Review Engine
================================
Streamlit web app for HVAC drawing analysis.
Upload a DXF/DWG drawing + BOQ spreadsheet â get a risk report.

Built by TechTelligence | nicholas@ttelligence.com
"""

import streamlit as st
import json
import os
import re
import io
import tempfile
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import the TraceQ engine (same directory)
from traceq_engine import TraceQEngine, Config, QuickScanResult, FileConverter


# âââ BOQ Parser âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ

# âââ BOQ Keyword Map ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
# Mapping: keywords in BOQ descriptions â engine equipment types
# Order matters â more specific matches first
# Tier 1 Synonym Library: mined from S1-S6 BOQ descriptions (May 25, 2026)
# 35 original keywords â 80+ to cover real-world BOQ description variations
BOQ_KEYWORD_MAP = [
    # === FCU variants (most specific first) ===
    ('FCU-1', 'fcu', 'FCU-1 Ducted'),
    ('FCU-2', 'fcu', 'FCU-2 Ducted'),
    ('FCU-3', 'fcu', 'FCU-3 Ducted'),
    ('FCU-4', 'fcu', 'FCU-4'),
    ('FCU', 'fcu', 'FCU (General)'),
    ('FAN COIL', 'fcu', 'Fan Coil Unit'),

    # === Thermostat ===
    ('THERMOSTAT', 'thermostat', 'Thermostat'),

    # === Flow bar / linear slot diffusers (BEFORE general diffusers) ===
    ('SUPPLY AIR FLOW BAR', 'flow_bar', 'Supply Air Flow Bar'),
    ('RETURN AIR FLOW BAR', 'flow_bar', 'Return Air Flow Bar'),
    ('FLOW BAR', 'flow_bar', 'Flow Bar'),
    ('SALD', 'flow_bar', 'Supply Air Linear Diffuser'),              # S1 BOQ abbreviation
    ('RALD', 'flow_bar', 'Return Air Linear Diffuser'),              # S1 BOQ abbreviation
    ('LENIAR SLOT DIFFUSER', 'flow_bar', 'Linear Slot Diffuser'),   # S3 typo variant
    ('LINEAR SLOT DIFFUSER', 'flow_bar', 'Linear Slot Diffuser'),   # S1
    ('SLOT DIFFUSER', 'flow_bar', 'Slot Diffuser'),                 # S2/S3 catch-all

    # === Diffusers â supply / return / extract ===
    ('SUPPLY AIR DIFFUSER', 'supply_diffuser', 'Supply Air Diffuser'),
    ('RETURN AIR DIFFUSER', 'return_diffuser', 'Return Air Diffuser'),
    ('EXTRACT DIFFUSER', 'extract_diffuser', 'Extract Diffuser'),
    ('EXHAUST SQUARE DIFFUSER', 'extract_diffuser', 'Exhaust Diffuser'),  # S1
    ('EXHAUST AIR DIFFUSER', 'extract_diffuser', 'Exhaust Air Diffuser'), # S2
    ('EXHAUST DIFFUSER', 'extract_diffuser', 'Exhaust Diffuser'),
    ('DECORATIVE DIFFUSER', 'supply_diffuser', 'Decorative Diffuser'),    # S1
    ('SQUARE DIFFUSER', 'supply_diffuser', 'Square Diffuser'),            # S1
    ('CIRCULAR DIFFUSER', 'supply_diffuser', 'Circular Diffuser'),        # S4
    ('DISK VALVE', 'extract_diffuser', 'Disk Valve'),                     # S3
    ('DISC VALVE', 'extract_diffuser', 'Disc Valve'),                     # S2/S3

    # === Plenum box ===
    ('PLENUM BOX', 'plenum_box', 'Plenum Box'),

    # === Dampers (most specific first) ===
    ('SUPPLY AIR VOLUME DAMPER', 'volume_control_damper', 'Supply Air Volume Damper'),
    ('VOLUME CONTROL DAMPER', 'volume_control_damper', 'Volume Control Damper'),
    ('VOLUME DAMPER', 'volume_control_damper', 'Volume Control Damper'),
    ('VCD', 'volume_control_damper', 'VCD'),
    ('SMOKE FIRE DAMPER', 'fire_damper', 'Smoke Fire Damper'),            # S2/S3
    ('FIRE DAMPER', 'fire_damper', 'Fire Damper'),
    ('MOTORIZED DAMPER', 'motorized_damper', 'Motorized Damper'),
    ('MOTORIZE DAMPER', 'motorized_damper', 'Motorize Damper'),           # S1 variant (no D)
    ('SMOKE DAMPER', 'fire_damper', 'Smoke Damper'),
    ('NON-RETURN DAMPER', 'non_return_damper', 'Non-Return Damper'),      # Hyphenated (S1)
    ('NON RETURN DAMPER', 'non_return_damper', 'Non-Return Damper'),
    ('BACK DRAFT DAMPER', 'non_return_damper', 'Backdraft Damper'),       # S2/S3

    # === Sound attenuator ===
    ('SOUND ATTENUATOR', 'sound_attenuator', 'Sound Attenuator'),
    ('ACOUSTIC', 'sound_attenuator', 'Acoustic Attenuator'),             # S1 "Acoustic linear"

    # === Ducts ===
    ('SUPPLY AIR DUCT', 'supply_duct', 'Supply Air Duct'),
    ('RETURN AIR DUCT', 'return_duct', 'Return Air Duct'),
    ('FLEXIBLE DUCT', 'flexible_duct', 'Flexible Duct'),

    # === VRF/VRV system â specific model patterns first, then units, then general ===
    ('VRV-IDU', 'indoor_unit', 'VRV Indoor Unit'),                       # S1 individual units
    ('DX-IDU', 'indoor_unit', 'DX Indoor Unit'),                         # S1 DX splits
    ('VRV-ODU', 'outdoor_unit', 'VRV Outdoor Unit'),                     # S1
    ('VRV-AHU', 'fahu', 'VRV-AHU Unit'),                                 # S1
    ('OUTDOOR UNIT', 'outdoor_unit', 'Outdoor Unit'),                    # BEFORE VRV/VRF (S5/S6)
    ('INDOOR UNIT', 'indoor_unit', 'Indoor Unit'),                       # BEFORE VRV/VRF
    ('VRV', 'vrf', 'VRV/VRF Unit'),
    ('VRF', 'vrf', 'VRF Unit'),
    ('DUCTED SPLIT', 'indoor_unit', 'Ducted Split Unit'),                # S2/S3
    ('DECORATIVE SPLIT', 'indoor_unit', 'Decorative Split Unit'),        # S3
    ('WALL MOUNTED', 'indoor_unit', 'Wall Mounted Unit'),

    # === AHU / FAHU ===
    ('FRESH AIR HANDLING', 'fahu', 'Fresh Air Handling Unit'),            # S2
    ('FAHU', 'fahu', 'FAHU'),
    ('AIR HANDLING UNIT', 'air_handling_unit', 'Air Handling Unit'),      # S3
    ('AHU', 'air_handling_unit', 'AHU'),

    # === Grilles & Louvers ===
    ('LINEAR BAR GRILLE', 'grille', 'Linear Bar Grille'),                # S1
    ('SLBG', 'grille', 'Supply Linear Bar Grille'),                      # S1 BOQ abbreviation
    ('GRILLE', 'grille', 'Grille'),
    ('SAND TRAP LOUVER', 'louver', 'Sand Trap Louver'),                  # S2/S3
    ('TRAP LOUVER', 'louver', 'Trap Louver'),                            # S1
    ('LOUVER', 'louver', 'Louver'),
    ('LOUVRE', 'louver', 'Louvre'),                                       # British spelling

    # === Fans ===
    ('EXHAUST FAN', 'exhaust_fan', 'Exhaust Fan'),
    ('EXTRACT FAN', 'exhaust_fan', 'Extract Fan'),                       # S2/S3
    ('VENTILATION FAN', 'exhaust_fan', 'Ventilation Fan'),
    ('SMOKE EXTRACT FAN', 'exhaust_fan', 'Smoke Extract Fan'),           # S2
    ('INLINE FAN', 'exhaust_fan', 'Inline Fan'),
    ('EAF-', 'exhaust_fan', 'Exhaust Air Fan'),                          # S1 EAF-01 etc
    ('FAF-', 'exhaust_fan', 'Fresh Air Fan'),                            # S1 FAF-04 etc

    # === Air curtain ===
    ('AIR CURTAIN', 'air_curtain', 'Air Curtain'),

    # === Miscellaneous ===
    ('ACCESS DOOR', 'access_door', 'Access Door'),
    ('CONDENSATE DRAIN', 'drain_pipe', 'Condensate Drain'),              # S3
    ('DRAIN', 'drain_pipe', 'Drain Pipe'),
    ('INSULATION', 'insulation', 'Insulation'),
]

# Units that can be compared directly (countable items)
COUNTABLE_UNITS = {'nos.', 'nos', 'no.', 'no', 'pcs', 'pcs.', 'ea', 'ea.', 'each', 'set', 'sets'}

# Trace ID prefix map: equipment_type â category prefix for TQ-[CAT]-[NNN] format
TRACE_PREFIX_MAP = {
    'supply_duct': 'DUCT', 'return_duct': 'DUCT', 'exhaust_duct': 'DUCT',
    'volume_control_damper': 'VCD', 'fcu': 'FCU',
    'supply_diffuser': 'DIFF', 'return_diffuser': 'DIFF', 'extract_diffuser': 'DIFF',
    'flow_bar': 'FLOW', 'thermostat': 'THERM', 'vrf': 'VRF',
    'flexible_duct': 'FLEX', 'plenum_box': 'PLEN',
    'drain_pipe': 'PIPE', 'refrigerant_pipe': 'PIPE',
    'sound_attenuator': 'ACOU', 'fire_damper': 'DAMP', 'motorized_damper': 'DAMP',
    'non_return_damper': 'DAMP', 'indoor_unit': 'FCU', 'outdoor_unit': 'VRF',
    'exhaust_fan': 'FAN', 'grille': 'DIFF', 'insulation': 'MISC',
    'access_door': 'MISC', 'damper_general': 'DAMP', 'hvac_equipment': 'EQUIP',
}

# Estimated UAE HVAC unit rates (AED) for missing-from-BOQ exposure calculation
UAE_UNIT_RATES = {
    'fcu': 2500, 'supply_diffuser': 180, 'return_diffuser': 180,
    'extract_diffuser': 200, 'volume_control_damper': 350, 'thermostat': 150,
    'vrf': 45000, 'flow_bar': 120, 'plenum_box': 280, 'indoor_unit': 2500,
    'outdoor_unit': 55000, 'fire_damper': 450, 'motorized_damper': 800,
    'sound_attenuator': 600, 'exhaust_fan': 3500, 'flexible_duct': 85,
}

# Expected detection method per equipment type â used in Trace ID reference
# when engine didn't detect the item (source = 'â')
EXPECTED_DETECTION_METHOD = {
    'supply_duct': 'Layer Detection', 'return_duct': 'Layer Detection', 'exhaust_duct': 'Layer Detection',
    'volume_control_damper': 'Layer Detection', 'fcu': 'Block Detection',
    'supply_diffuser': 'Layer Detection', 'return_diffuser': 'Block Detection',
    'extract_diffuser': 'Text/Label Detection', 'thermostat': 'Layer Detection',
    'flow_bar': 'Text/Label Detection', 'vrf': 'Text/Label Detection',
    'flexible_duct': 'Layer Detection', 'plenum_box': 'Text/Label Detection',
    'indoor_unit': 'Block Detection', 'outdoor_unit': 'Block Detection',
    'fire_damper': 'Block Detection', 'motorized_damper': 'Block Detection',
    'non_return_damper': 'Block Detection', 'sound_attenuator': 'Block Detection',
    'exhaust_fan': 'Block Detection', 'grille': 'Block Detection',
    'insulation': 'Layer Detection', 'access_door': 'Block Detection',
    'drain_pipe': 'Layer Detection', 'refrigerant_pipe': 'Layer Detection',
}


def _detect_boq_columns(ws):
    """
    Detect which columns hold description, unit, qty, rate, total.
    Scans first 10 rows for header keywords, then validates against actual data.
    Returns dict with col indices (1-based).
    """
    cols = {'desc': 3, 'unit': 4, 'qty': 5, 'rate': 6, 'total': 7, 'item_no': 2}

    for row_num in range(1, min(11, ws.max_row + 1)):
        for col_num in range(1, min(11, ws.max_column + 1)):
            val = ws.cell(row=row_num, column=col_num).value
            if val is None:
                continue
            upper = str(val).strip().upper()
            if upper in ('UNIT', 'UOM', 'U/M'):
                cols['unit'] = col_num
            elif upper in ('QTY', 'QUANTITY', 'QTY.'):
                cols['qty'] = col_num
            elif 'RATE' in upper and 'TOTAL' not in upper:
                cols['rate'] = col_num
            elif 'TOTAL' in upper:
                cols['total'] = col_num

    text_col_counts = {}
    for row_num in range(5, min(20, ws.max_row + 1)):
        for col_num in range(1, min(8, ws.max_column + 1)):
            val = ws.cell(row=row_num, column=col_num).value
            if isinstance(val, str) and len(val.strip()) > 5:
                text_col_counts[col_num] = text_col_counts.get(col_num, 0) + 1

    if text_col_counts:
        best_col = max(text_col_counts, key=text_col_counts.get)
        cols['desc'] = best_col
        cols['item_no'] = best_col - 1 if best_col > 1 else 1

    return cols


def _classify_description(desc_text):
    upper = desc_text.upper().strip()
    for keyword, etype, label in BOQ_KEYWORD_MAP:
        if keyword in upper:
            return etype, label
    return None, desc_text



def boq_coverage_check(boq_items):
    """
    BOQ Coverage Check (pre-flight).
    Cross-references every BOQ line item against the keyword map classification.
    Returns (classified, unclassified) â lists of dicts with item info.
    Tier 1 pre-flight: flags items that can\'t be matched to any equipment type
    so the user knows BEFORE the report generates what fell through.
    """
    classified = []
    unclassified = []

    for item in boq_items:
        desc = item.get('description', '')
        etype = item.get('equipment_type')
        label = item.get('equipment_label', '')
        qty = item.get('qty', 0)

        entry = {
            'description': desc[:80] + ('...' if len(desc) > 80 else ''),
            'full_description': desc,
            'qty': qty,
            'unit': item.get('unit', ''),
            'equipment_type': etype,
            'equipment_label': label,
        }

        if etype is not None:
            classified.append(entry)
        else:
            unclassified.append(entry)

    return classified, unclassified


def parse_boq(file_bytes, filename):
    """
    Parse a BOQ Excel file and extract equipment line items.
    Uses column-position detection (not heuristic scanning).
    """
    suffix = os.path.splitext(filename)[1].lower() or '.xlsx'
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name

    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        ws = wb.active

        cols = _detect_boq_columns(ws)
        items = []
        item_counter = 0

        for row_num in range(1, ws.max_row + 1):
            item_no_val = ws.cell(row=row_num, column=cols['item_no']).value
            desc_val = ws.cell(row=row_num, column=cols['desc']).value
            unit_val = ws.cell(row=row_num, column=cols['unit']).value
            qty_val = ws.cell(row=row_num, column=cols['qty']).value
            rate_val = ws.cell(row=row_num, column=cols['rate']).value
            total_val = ws.cell(row=row_num, column=cols['total']).value

            if desc_val is None:
                continue
            desc_str = str(desc_val).strip()
            if len(desc_str) < 3:
                continue
            if qty_val is None:
                continue
            try:
                qty = float(qty_val)
            except (ValueError, TypeError):
                continue
            if qty <= 0:
                continue

            equip_type, equip_label = _classify_description(desc_str)
            unit_str = str(unit_val).strip() if unit_val else None
            try:
                rate = float(rate_val) if rate_val else None
            except (ValueError, TypeError):
                rate = None
            try:
                total = float(total_val) if total_val else None
            except (ValueError, TypeError):
                total = None

            item_counter += 1
            boq_ref = str(item_no_val).strip() if item_no_val else str(item_counter)

            items.append({
                'description': desc_str,
                'equipment_type': equip_type,
                'equipment_label': equip_label,
                'qty': qty,
                'unit': unit_str,
                'rate': rate,
                'total': total,
                'boq_ref': boq_ref,
                'is_countable': (unit_str or '').lower().strip('.') in {'nos', 'no', 'pcs', 'ea', 'each', 'set', 'sets'},
            })

        os.unlink(tmp_path)
        return items

    except Exception as e:
        os.unlink(tmp_path)
        raise e


def compare_boq_vs_drawing(boq_items, drawing_merged):
    """
    Compare BOQ line items against drawing detection results.
    Returns (comparisons, missing_from_boq) with category-based Trace IDs.
    Status labels: MATCH, DISCREPANCY only (no risk levels).
    Items appear in BOQ order (first-seen equipment type order).
    """
    comparisons = []
    trace_counters = {}  # per-category counters for TQ-[CAT]-[NNN]

    def _make_trace_id(etype):
        prefix = TRACE_PREFIX_MAP.get(etype, 'EQUIP')
        trace_counters[prefix] = trace_counters.get(prefix, 0) + 1
        return f"TQ-{prefix}-{trace_counters[prefix]:03d}"

    # Group BOQ items by equipment type, preserving first-seen order
    boq_by_type = {}
    boq_type_order = []  # preserves BOQ order
    for item in boq_items:
        etype = item['equipment_type']
        if etype is None:
            continue
        if etype not in boq_by_type:
            boq_by_type[etype] = {
                'total_qty': 0,
                'items': [],
                'total_cost': 0,
                'rates': [],
                'units': set(),
            }
            boq_type_order.append(etype)
        boq_by_type[etype]['total_qty'] += item['qty']
        boq_by_type[etype]['items'].append(item)
        if item.get('total'):
            boq_by_type[etype]['total_cost'] += item['total']
        if item.get('rate') and item['rate'] > 0:
            boq_by_type[etype]['rates'].append(item['rate'])
        if item.get('unit'):
            boq_by_type[etype]['units'].add(item['unit'].strip().lower())

    # Build comparison for each BOQ equipment type â IN BOQ ORDER
    matched_drawing_types = set()

    for etype in boq_type_order:
        boq_data = boq_by_type[etype]
        drawing_data = drawing_merged.get(etype, {})
        matched_drawing_types.add(etype)

        boq_qty = boq_data['total_qty']
        drawing_qty = drawing_data.get('count', 0)
        source = drawing_data.get('source', 'â')
        rates = boq_data['rates']
        avg_rate = sum(rates) / len(rates) if rates else UAE_UNIT_RATES.get(etype, 0)
        units = boq_data['units']

        has_non_countable = any(u.strip('.') not in ('nos', 'no', 'pcs', 'ea', 'each', 'set', 'sets') for u in units)

        diff = drawing_qty - boq_qty
        exposure = abs(diff) * avg_rate if avg_rate and diff != 0 else 0

        # Determine status â ONLY MATCH or DISCREPANCY (no risk levels)
        # 0% tolerance: only exact match (diff == 0 AND drawing actually detected) = MATCH
        # If drawing_qty == 0 and boq_qty > 0, that's NOT a match â it's not detected
        is_unit_mismatch = has_non_countable
        if diff == 0 and drawing_qty > 0 and not is_unit_mismatch:
            status = 'MATCH'
            note = "Exact match."
        else:
            status = 'DISCREPANCY'
            if is_unit_mismatch:
                note = _build_unit_mismatch_note(etype, boq_data, drawing_data, units)
            elif drawing_qty == 0:
                note = f"BOQ has {int(boq_qty)} but not detected in drawing. Manual verification required."
            else:
                note = _build_discrepancy_note(etype, boq_data, drawing_data, diff)

        boq_breakdown = _format_boq_breakdown(boq_data['items'])
        name = _format_equipment_name(etype)
        source_label = _format_source_label(source)
        trace_id = _make_trace_id(etype)

        # Always show drawing qty â even for unit mismatches, show entity count
        if drawing_qty > 0:
            if is_unit_mismatch:
                show_drawing_qty = f"{int(drawing_qty)} entities"
            else:
                show_drawing_qty = int(drawing_qty)
            show_diff = f"{int(diff):+d}" if not is_unit_mismatch else 'â'
            variance_pct = f"{abs(diff) / max(boq_qty, 1) * 100:.0f}%" if not is_unit_mismatch and boq_qty > 0 else 'â'
        else:
            show_drawing_qty = 'Not Detected'
            show_diff = 'â'
            variance_pct = 'â'

        # Only show exposure for DISCREPANCY items (0% tolerance â exact match only)
        if status == 'DISCREPANCY' and exposure > 0 and not is_unit_mismatch:
            show_exposure = f"{exposure:,.0f}"
        else:
            show_exposure = 'â'
            if status == 'MATCH':
                exposure = 0  # zero out exposure for match items

        comparisons.append({
            'Trace ID': trace_id,
            'Equipment': name,
            'BOQ Qty': int(boq_qty) if boq_qty == int(boq_qty) else f"{boq_qty:,.1f}",
            'Drawing Qty': show_drawing_qty,
            'Difference': show_diff,
            'Variance %': variance_pct,
            'Unit': ', '.join(sorted(units)) if units else 'â',
            'Risk': status,  # field name kept for backwards compat; values are MATCH/DISCREPANCY only
            'Exposure (AED)': show_exposure,
            'Notes': note,
            'BOQ Breakdown': boq_breakdown,
            'Detection Source': source_label,
            '_exposure_num': exposure if not is_unit_mismatch else 0,
            '_boq_qty': boq_qty,
            '_drawing_qty': drawing_qty if drawing_qty > 0 else None,
            '_diff': diff if not is_unit_mismatch else None,
            '_rate': avg_rate,
            '_is_unit_mismatch': is_unit_mismatch,
            '_equipment_type': etype,
        })

    # Missing from BOQ: items in drawing but not in any BOQ type
    missing_from_boq = []
    for etype, data in sorted(drawing_merged.items()):
        if etype not in matched_drawing_types and data.get('count', 0) > 0:
            trace_id = _make_trace_id(etype) if etype in TRACE_PREFIX_MAP else f"TQ-MISS-{trace_counters.get('MISS', 0) + 1:03d}"
            if etype not in TRACE_PREFIX_MAP:
                trace_counters['MISS'] = trace_counters.get('MISS', 0) + 1
            name = _format_equipment_name(etype)
            source = data.get('source', 'unknown')
            source_label = _format_source_label(source)
            confidence = data.get('confidence', 0)
            qty = data['count']
            unit_rate = UAE_UNIT_RATES.get(etype, 0)
            est_exposure = qty * unit_rate

            missing_from_boq.append({
                'Trace ID': trace_id,
                'Equipment': name,
                'Drawing Qty': qty,
                'Detection': source_label,
                'Confidence': f"{int(confidence * 100)}%",
                'Notes': f"Found in drawing via {source_label.lower()} but no matching BOQ line item.",
                '_equipment_type': etype,
                '_unit_rate': unit_rate,
                '_est_exposure': est_exposure,
            })

    return comparisons, missing_from_boq


def _format_equipment_name(etype):
    """Format equipment type to proper display name, preserving acronyms."""
    acronyms = {'fcu': 'FCU', 'vrf': 'VRF', 'vcd': 'VCD', 'sad': 'SAD', 'rad': 'RAD'}
    name = etype.replace('_', ' ').title()
    for key, acr in acronyms.items():
        name = name.replace(key.title(), acr)
    return name


def _format_source_label(source):
    """Convert raw engine source to clean label."""
    if not source or source == 'â':
        return 'â'
    if 'tier1' in source:
        return 'Layer Detection'
    elif 'tier2' in source:
        return 'Block Detection'
    elif 'tier3' in source or 'SAD' in source:
        return 'Text/Label Detection'
    return source


def _format_boq_breakdown(items):
    if len(items) == 1:
        return items[0]['description'][:60]
    parts = []
    for item in items:
        desc_short = item['description'][:40]
        parts.append(f"{desc_short}: {int(item['qty']) if item['qty'] == int(item['qty']) else item['qty']}")
    return " | ".join(parts)


def _build_discrepancy_note(etype, boq_data, drawing_data, diff):
    drawing_qty = drawing_data.get('count', 0)
    boq_qty = boq_data['total_qty']
    source = drawing_data.get('source', 'detection')

    if 'tier1' in source:
        source_label = "layer detection"
    elif 'tier2' in source:
        source_label = "block detection"
    elif 'tier3' in source or 'SAD' in source:
        source_label = "text/label detection"
    else:
        source_label = source

    if diff > 0:
        note = f"Drawing shows {int(drawing_qty)} via {source_label} â {abs(int(diff))} more than BOQ ({int(boq_qty)})."
    else:
        note = f"Drawing shows {int(drawing_qty)} via {source_label} â {abs(int(diff))} fewer than BOQ ({int(boq_qty)})."

    if len(boq_data['items']) > 1:
        sub_parts = []
        for item in boq_data['items']:
            short = item['description'][:35]
            sub_parts.append(f"{short}={int(item['qty'])}")
        note += f" BOQ breakdown: {', '.join(sub_parts)}."

    if etype == 'return_diffuser' and diff > 0:
        note += " Note: block detection may double-count *U16/*U17 inserts â verify against drawing legend."
    elif etype == 'vrf' and diff < 0:
        note += " Note: drawing detects unique VRF labels only â BOQ may list individual modules per system."
    elif etype == 'flow_bar' and diff < 0:
        note += " Note: flow bars often lack distinct block markers â text detection may undercount."
    elif etype == 'volume_control_damper' and len(boq_data['items']) > 1:
        note += " Note: BOQ may list different damper sizes separately â verify each size against drawing."

    return note


def _build_unit_mismatch_note(etype, boq_data, drawing_data, units):
    """Build note for items where BOQ units differ from TraceQ entity count."""
    unit_list = ', '.join(sorted(units))
    boq_qty = boq_data['total_qty']
    drawing_qty = drawing_data.get('count', 0)
    source = drawing_data.get('source', 'detection')
    source_label = _format_source_label(source).lower()

    if drawing_qty > 0:
        base = f"BOQ = {boq_qty:,.1f} {unit_list}. TraceQ found {int(drawing_qty)} entities via {source_label}."
    else:
        base = f"BOQ = {boq_qty:,.1f} {unit_list}. Not detected in drawing."

    if any(u.strip('.') in ('sqm', 'sq.m', 'sq m', 'm2', 'sqft') for u in units):
        base += " BOQ measured in area â duct schedule comparison recommended."
    elif any(u.strip('.') in ('mtrs', 'mtr', 'm', 'lm', 'rm') for u in units):
        base += " BOQ measured in length â direct entity comparison not possible."

    return base


def _build_verify_note(etype, boq_data, units):
    """Legacy function â kept for backwards compatibility."""
    return _build_unit_mismatch_note(etype, boq_data, {}, units)


# âââ Excel Report Generator ââââââââââââââââââââââââââââââââââââââââââââââââââ

def _xl_val(v):
    """Return 'â' string for None values in Excel cells."""
    return 'â' if v is None else v


def generate_validator_template(comparisons, missing_from_boq, merged, drawing_name, scan=None):
    """
    Generate a 4-tab validator XLSX for QS verification.
    Tab 1: Instructions â what to do, time budget, rules
    Tab 2: BOQ Comparison â engine output (locked) + Agree? (Yes/No/Partial) + Comments
    Tab 3: Missing from BOQ â engine findings + Agree? + Comments
    Tab 4: Validator Discoveries â blank rows for items the QS finds that engine missed
    Returns bytes of the .xlsx file.
    """
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()

    # ââ Shared styles ââ
    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2C3E50')
    section_font = Font(name='Arial', bold=True, size=10, color='8B0000')
    section_fill = PatternFill('solid', fgColor='F5F0EB')
    data_font = Font(name='Arial', size=10)
    yellow_fill = PatternFill('solid', fgColor='FFFDE7')
    light_gray_fill = PatternFill('solid', fgColor='F8F8F8')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )
    wrap_align = Alignment(wrap_text=True, vertical='top')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    drawing_short = drawing_name.replace('.dxf', '').replace('.DXF', '').replace('.dwg', '').replace('.DWG', '')
    if len(drawing_short) > 80:
        file_count = drawing_short.count(' + ') + 1
        first_file = drawing_short.split(' + ')[0]
        drawing_short = f"{first_file} + {file_count - 1} more files"

    def _style_header_row(ws, row, max_col):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

    def _style_data_row(ws, row, max_col, is_alt=False, yellow_cols=None):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = wrap_align
            if yellow_cols and col in yellow_cols:
                cell.fill = yellow_fill
            elif is_alt:
                cell.fill = light_gray_fill

    def _add_section_label(ws, row, col, text, max_col):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = section_font
        cell.fill = section_fill
        for c in range(col, max_col + 1):
            ws.cell(row=row, column=c).fill = section_fill
            ws.cell(row=row, column=c).border = thin_border

    # âââ TAB 1: INSTRUCTIONS âââ
    ws1 = wb.active
    ws1.title = 'Instructions'
    ws1.sheet_properties.tabColor = '2C3E50'
    ws1.column_dimensions['A'].width = 80

    instr_rows = [
        ('TraceQ Validator Brief', Font(name='Arial', bold=True, size=14, color='8B0000')),
        ('', None),
        ('Your job in one sentence:', Font(name='Arial', bold=True, size=11)),
        ("Verify the engine's findings line by line and flag anything the engine got wrong or missed.", data_font),
        ('', None),
        (f'Drawing: {drawing_short}', Font(name='Arial', bold=True, size=11)),
        (f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}', data_font),
        ('', None),
        ('TIME BUDGET', Font(name='Arial', bold=True, size=11)),
        ('Target: 2-3 hours total. This is verification work, not a full takeoff.', data_font),
        ('', None),
        ('WHAT TO DO', Font(name='Arial', bold=True, size=11)),
        ('1. BOQ Comparison tab: For each row, check whether the engine count is correct.', data_font),
        ('   - Yes = engine count is right (or close enough to not matter commercially)', data_font),
        ('   - No = engine count is wrong (explain why in Comments column)', data_font),
        ('   - Partial = engine is close but needs adjustment (note the correct count in Comments)', data_font),
        ('', None),
        ('2. Items Missing from BOQ tab: Verify each item the engine found on drawings but not in the BOQ.', data_font),
        ('   - Yes = the item genuinely exists on drawings and is missing from the BOQ', data_font),
        ('   - No = the engine misidentified something (explain what it actually is in Comments)', data_font),
        ('', None),
        ('3. Validator Discoveries tab: Add any items YOU found that the engine missed entirely.', data_font),
        ('   - If the engine caught everything, leave this sheet blank.', data_font),
        ('   - These discoveries help us improve the engine for future jobs.', data_font),
        ('', None),
        ('WHAT NOT TO DO', Font(name='Arial', bold=True, size=11)),
        ('- Do not recount from scratch. Spot-check the engine output against the drawings.', data_font),
        ('- Do not edit the engine output columns (white background). Only fill the yellow cells.', data_font),
        ('- Do not write client-facing prose. Short technical notes are fine.', data_font),
    ]
    for i, (text, font) in enumerate(instr_rows, 1):
        cell = ws1.cell(row=i, column=1, value=text)
        if font:
            cell.font = font

    # âââ TAB 2: BOQ COMPARISON âââ
    ws2 = wb.create_sheet('BOQ Comparison')
    ws2.sheet_properties.tabColor = '27AE60'

    _add_section_label(ws2, 1, 1, "Engine output (do not edit)", 8)
    _add_section_label(ws2, 1, 7, "Validator (fill the yellow cells)", 8)

    headers2 = ['#', 'Equipment (from BOQ)', 'Unit', 'BOQ Qty', 'Engine Count',
                'Variance', 'Agree?', 'Comments']
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=2, column=col, value=h)
    _style_header_row(ws2, 2, 8)

    row = 3
    for i, comp in enumerate(comparisons, 1):
        equip = comp['Equipment']
        boq_qty = comp.get('_boq_qty', 0)
        dwg_qty = comp.get('_drawing_qty', 0)
        unit = comp.get('Unit', 'nos.')
        is_unit_mismatch = comp.get('_is_unit_mismatch', False)

        # Calculate variance
        if is_unit_mismatch:
            variance = 'Unit mismatch'
        elif dwg_qty in (None, 'â', 0, '') or dwg_qty == 0:
            variance = 'NOT DETECTED'
        else:
            try:
                diff = float(dwg_qty) - float(boq_qty)
                pct = diff / max(float(boq_qty), 1) * 100
                variance = f'{diff:+.0f} ({pct:+.1f}%)'
            except (ValueError, TypeError):
                variance = 'â'

        ws2.cell(row=row, column=1, value=i)
        ws2.cell(row=row, column=2, value=equip)
        ws2.cell(row=row, column=3, value=unit)
        ws2.cell(row=row, column=4, value=boq_qty)
        ws2.cell(row=row, column=5, value=dwg_qty if dwg_qty not in (None, '') else 0)
        ws2.cell(row=row, column=6, value=variance)
        ws2.cell(row=row, column=7, value='')  # Agree?
        ws2.cell(row=row, column=8, value='')  # Comments

        _style_data_row(ws2, row, 8, is_alt=(i % 2 == 0), yellow_cols={7, 8})
        row += 1

    # Data validation for Agree? column
    dv2 = DataValidation(type="list", formula1='"Yes,No,Partial"', allow_blank=True)
    dv2.error = "Please select: Yes, No, or Partial"
    dv2.errorTitle = "Invalid Entry"
    ws2.add_data_validation(dv2)
    if row > 3:
        dv2.add(f'G3:G{row - 1}')

    widths2 = [5, 30, 8, 10, 12, 18, 12, 40]
    for col, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # âââ TAB 3: MISSING FROM BOQ âââ
    ws3 = wb.create_sheet('Missing from BOQ')
    ws3.sheet_properties.tabColor = 'E74C3C'

    _add_section_label(ws3, 1, 1, "Engine output", 7)
    _add_section_label(ws3, 1, 6, "Validator (fill the yellow cells)", 7)

    headers3 = ['#', 'Equipment Description', 'Unit', 'Engine Count',
                'Detection Source', 'Agree?', 'Comments']
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=2, column=col, value=h)
    _style_header_row(ws3, 2, 7)

    miss_row = 3
    miss_count = 0
    for m in missing_from_boq:
        miss_count += 1
        qty = m.get('Drawing Qty', 0)
        detection = m.get('Detection', '?')

        ws3.cell(row=miss_row, column=1, value=miss_count)
        ws3.cell(row=miss_row, column=2, value=m['Equipment'])
        ws3.cell(row=miss_row, column=3, value='nos.')
        ws3.cell(row=miss_row, column=4, value=qty)
        ws3.cell(row=miss_row, column=5, value=detection)
        ws3.cell(row=miss_row, column=6, value='')  # Agree?
        ws3.cell(row=miss_row, column=7, value='')  # Comments

        _style_data_row(ws3, miss_row, 7, is_alt=(miss_count % 2 == 0), yellow_cols={6, 7})
        miss_row += 1

    if miss_count == 0:
        ws3.cell(row=3, column=1, value='No additional items found by engine beyond BOQ categories.')
        ws3.cell(row=3, column=1).font = Font(name='Arial', italic=True, size=10, color='888888')

    # Data validation for Agree? column
    dv3 = DataValidation(type="list", formula1='"Yes,No,Partial"', allow_blank=True)
    ws3.add_data_validation(dv3)
    if miss_count > 0:
        dv3.add(f'F3:F{miss_row - 1}')

    widths3 = [5, 28, 8, 12, 25, 12, 40]
    for col, w in enumerate(widths3, 1):
        ws3.column_dimensions[get_column_letter(col)].width = w

    # âââ TAB 4: VALIDATOR DISCOVERIES âââ
    ws4 = wb.create_sheet('Validator Discoveries')
    ws4.sheet_properties.tabColor = '8E44AD'

    ws4.cell(row=1, column=1, value="Items YOU found on the drawings that the engine MISSED entirely")
    ws4.cell(row=1, column=1).font = Font(name='Arial', bold=True, size=11, color='8B0000')
    ws4.cell(row=2, column=1, value="If everything is already covered, leave this sheet blank. We use this to improve the engine.")
    ws4.cell(row=2, column=1).font = Font(name='Arial', italic=True, size=10, color='666666')

    headers4 = ['#', 'Equipment Description', 'Unit', 'Your Count',
                'Where in drawing (floor/area)', 'Why you think the engine missed it']
    for col, h in enumerate(headers4, 1):
        ws4.cell(row=3, column=col, value=h)
    _style_header_row(ws4, 3, 6)

    for r in range(4, 19):
        ws4.cell(row=r, column=1, value=r - 3)
        _style_data_row(ws4, r, 6, yellow_cols={2, 3, 4, 5, 6})

    widths4 = [5, 28, 8, 12, 30, 40]
    for col, w in enumerate(widths4, 1):
        ws4.column_dimensions[get_column_letter(col)].width = w

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_excel_report(comparisons, missing_from_boq, boq_items, drawing_name, boq_name, merged=None, dedup_report=None, validation_metadata=None):
    """
    Generate a professional Excel BOQ Risk Analysis Report per TraceQ_Report_Format_Spec.
    3-tab client report:
      Tab 1: Executive Summary (stats bar, top findings, total AED exposure)
      Tab 2: BOQ Comparison (BOQ order, all items, AED totals row)
      Tab 3: Missing from BOQ (context one-liner, exposure, disclaimer)
    Returns bytes of the .xlsx file.
    """
    wb = openpyxl.Workbook()

    # âââ STYLES (per Format Spec) âââââââââââââââââââââââââââââââââââââââââââ
    navy = '002060'
    dark_blue = '1F4E79'
    col_header_blue = '4472C4'
    section_bg = 'D6E4F0'
    match_bg = 'E2EFDA'
    match_text = '375623'
    discrep_bg = 'FCE4EC'
    discrep_text = 'C62828'
    missing_bg = 'FFF3E0'
    missing_text = 'E65100'
    alt_row_bg = 'F2F2F2'
    border_color = 'D9D9D9'

    banner_font = Font(name='Arial', bold=True, size=18, color='FFFFFF')
    sub_banner_font = Font(name='Arial', italic=True, size=10, color='FFFFFF')
    title_font = Font(name='Arial', bold=True, size=14, color=navy)
    subtitle_font = Font(name='Arial', size=10, color='666666')
    bold_font = Font(name='Arial', bold=True, size=10)
    bold_font_big = Font(name='Arial', bold=True, size=12)
    normal_font = Font(name='Arial', size=10)
    notes_font = Font(name='Arial', size=9, color='555555')
    trace_font = Font(name='Arial', bold=True, size=8, color=dark_blue)
    aed_font = Font(name='Arial', bold=True, size=10, color=discrep_text)
    aed_big_font = Font(name='Arial', bold=True, size=16, color=discrep_text)
    disclaimer_font = Font(name='Arial', italic=True, size=8, color='888888')

    col_header_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    col_header_fill = PatternFill('solid', fgColor=col_header_blue)
    navy_fill = PatternFill('solid', fgColor=navy)
    dark_blue_fill = PatternFill('solid', fgColor=dark_blue)
    section_fill = PatternFill('solid', fgColor=section_bg)
    match_fill = PatternFill('solid', fgColor=match_bg)
    discrep_fill = PatternFill('solid', fgColor=discrep_bg)
    missing_fill = PatternFill('solid', fgColor=missing_bg)
    alt_fill = PatternFill('solid', fgColor=alt_row_bg)

    match_status_font = Font(name='Arial', bold=True, size=10, color=match_text)
    discrep_status_font = Font(name='Arial', bold=True, size=10, color=discrep_text)
    missing_status_font = Font(name='Arial', bold=True, size=10, color=missing_text)

    thin_border = Border(
        left=Side(style='thin', color=border_color),
        right=Side(style='thin', color=border_color),
        top=Side(style='thin', color=border_color),
        bottom=Side(style='thin', color=border_color),
    )

    def _apply_border_row(ws, row_num, max_col):
        for ci in range(1, max_col + 1):
            ws.cell(row=row_num, column=ci).border = thin_border

    now = datetime.now().strftime('%d %B %Y, %H:%M')

    # Truncate long multi-file drawing names
    drawing_display = drawing_name.replace('.dxf', '').replace('.DXF', '').replace('.dwg', '').replace('.DWG', '')
    if len(drawing_display) > 80:
        file_count = drawing_display.count(' + ') + 1
        first_file = drawing_display.split(' + ')[0]
        drawing_display = f"{first_file} + {file_count - 1} more files"

    # Pre-calculate stats
    matches = sum(1 for c in comparisons if c['Risk'] == 'MATCH')
    discrepancies = sum(1 for c in comparisons if c['Risk'] == 'DISCREPANCY')
    missing_count = len(missing_from_boq)
    total_items = len(comparisons)
    comparison_exposure = sum(c.get('_exposure_num', 0) or 0 for c in comparisons)
    missing_exposure = sum(m.get('_est_exposure', 0) or 0 for m in missing_from_boq)
    total_exposure = comparison_exposure + missing_exposure

    # âââ TAB 1: EXECUTIVE SUMMARY âââââââââââââââââââââââââââââââââââââââââââ
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.sheet_properties.tabColor = navy

    # Navy banner
    ws1.merge_cells('A1:F2')
    c = ws1['A1']
    c.value = 'TraceQ â HVAC BOQ Risk Analysis'
    c.font = banner_font
    c.fill = navy_fill
    c.alignment = Alignment(horizontal='center', vertical='center')
    for ci in range(1, 7):
        for ri in (1, 2):
            ws1.cell(row=ri, column=ci).fill = navy_fill

    # Sub-banner
    ws1.merge_cells('A3:F3')
    c = ws1['A3']
    c.value = f'Prepared by TechTelligence | {now}'
    c.font = sub_banner_font
    c.fill = dark_blue_fill
    c.alignment = Alignment(horizontal='center')
    for ci in range(1, 7):
        ws1.cell(row=3, column=ci).fill = dark_blue_fill

    # Project details
    details = [
        ('Drawing:', drawing_display),
        ('BOQ:', boq_name),
        ('Analysis Date:', now),
    ]
    row = 5
    for label, val in details:
        ws1.cell(row=row, column=1, value=label).font = bold_font
        ws1.cell(row=row, column=2, value=val).font = normal_font
        row += 1

    # ── Validation metadata (if provided) ──
    if validation_metadata:
        row += 1
        c = ws1.cell(row=row, column=1, value='VALIDATION SUMMARY')
        c.font = Font(name='Arial', bold=True, size=11, color=navy)
        for ci in range(1, 7):
            ws1.cell(row=row, column=ci).fill = section_fill
            ws1.cell(row=row, column=ci).border = thin_border
        row += 1
        vm = validation_metadata
        val_details = [
            ('Method:', vm.get('validation_method', 'N/A')),
            ('Validator(s):', ', '.join(vm.get('validator_names', []))),
            ('Engine Errors Excluded:', str(vm.get('engine_errors_excluded', 0))),
            ('Validator Corrections:', str(vm.get('validator_corrections', 0))),
            ('Discoveries Added:', str(vm.get('discoveries_added', 0))),
        ]
        if vm.get('conservative_includes', 0) > 0:
            val_details.append(('Disagreements (conservatively included):', str(vm['conservative_includes'])))
        for label, val in val_details:
            ws1.cell(row=row, column=1, value=label).font = bold_font
            ws1.cell(row=row, column=2, value=val).font = normal_font
            _apply_border_row(ws1, row, 2)
            row += 1

    # ââ Stats Bar ââ
    row += 1
    stats_headers = ['Total Items Reviewed', 'Matched', 'Discrepancies', 'Missing from BOQ', 'Total Est. Exposure (AED)']
    stats_values = [total_items, matches, discrepancies, missing_count, total_exposure]

    for col_idx, h in enumerate(stats_headers, 1):
        c = ws1.cell(row=row, column=col_idx, value=h)
        c.font = col_header_font
        c.fill = col_header_fill
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = thin_border
    row += 1
    for col_idx, v in enumerate(stats_values, 1):
        c = ws1.cell(row=row, column=col_idx, value=v)
        c.font = bold_font_big if col_idx < 5 else aed_font
        c.alignment = Alignment(horizontal='center')
        c.border = thin_border
        if isinstance(v, (int, float)):
            c.number_format = '#,##0'
    row += 2

    # ââ Top Findings ââ
    c = ws1.cell(row=row, column=1, value='KEY FINDINGS')
    c.font = Font(name='Arial', bold=True, size=11, color=navy)
    for ci in range(1, 7):
        ws1.cell(row=row, column=ci).fill = section_fill
        ws1.cell(row=row, column=ci).border = thin_border
    row += 1

    finding_headers = ['Finding', 'Status', 'Detail']
    for col_idx, h in enumerate(finding_headers, 1):
        c = ws1.cell(row=row, column=col_idx, value=h)
        c.font = col_header_font
        c.fill = col_header_fill
        c.alignment = Alignment(horizontal='center')
        c.border = thin_border
    row += 1

    # Show discrepancy items + missing items as findings
    for comp in comparisons:
        if comp['Risk'] == 'MATCH':
            continue
        exp = comp.get('_exposure_num', 0) or 0
        detail = comp['Notes'][:120]
        if exp > 0:
            detail += f" Est. exposure: AED {exp:,.0f}"

        ws1.cell(row=row, column=1, value=comp['Equipment']).font = normal_font
        sc = ws1.cell(row=row, column=2, value='DISCREPANCY')
        sc.font = discrep_status_font
        sc.fill = discrep_fill
        ws1.cell(row=row, column=3, value=detail).font = notes_font
        ws1.cell(row=row, column=3).alignment = Alignment(wrap_text=True, vertical='top')
        _apply_border_row(ws1, row, 3)
        row += 1

    for m in missing_from_boq:
        exp = m.get('_est_exposure', 0) or 0
        detail = f"{m['Drawing Qty']} found in drawing, not in BOQ."
        if exp > 0:
            detail += f" Est. exposure: AED {exp:,.0f}"

        ws1.cell(row=row, column=1, value=m['Equipment']).font = normal_font
        sc = ws1.cell(row=row, column=2, value='MISSING FROM BOQ')
        sc.font = missing_status_font
        sc.fill = missing_fill
        ws1.cell(row=row, column=3, value=detail).font = notes_font
        ws1.cell(row=row, column=3).alignment = Alignment(wrap_text=True, vertical='top')
        _apply_border_row(ws1, row, 3)
        row += 1

    # ââ Total Estimated Financial Exposure ââ
    row += 1
    ws1.cell(row=row, column=1, value='TOTAL ESTIMATED FINANCIAL EXPOSURE').font = bold_font_big
    row += 1
    c = ws1.cell(row=row, column=1, value=f'AED {total_exposure:,.0f}')
    c.font = aed_big_font
    row += 1
    breakdown = f'Discrepancies: AED {comparison_exposure:,.0f}  |  Missing items: AED {missing_exposure:,.0f}'
    ws1.cell(row=row, column=1, value=breakdown).font = normal_font
    row += 1
    ws1.cell(row=row, column=1, value='Estimated based on typical UAE HVAC market rates â indicative only.').font = disclaimer_font

    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 70
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 22
    ws1.column_dimensions['F'].width = 14

    # âââ TAB 2: BOQ COMPARISON ââââââââââââââââââââââââââââââââââââââââââââââ
    ws2 = wb.create_sheet("BOQ Comparison")
    ws2.sheet_properties.tabColor = col_header_blue

    # Banner
    ws2.merge_cells('A1:K1')
    c = ws2['A1']
    c.value = 'BOQ vs Drawing â Detailed Comparison'
    c.font = banner_font
    c.fill = navy_fill
    c.alignment = Alignment(horizontal='center', vertical='center')
    for ci in range(1, 12):
        ws2.cell(row=1, column=ci).fill = navy_fill

    ws2.merge_cells('A2:K2')
    c = ws2['A2']
    c.value = f'Drawing: {drawing_display}  |  BOQ: {boq_name}  |  Date: {now}'
    c.font = sub_banner_font
    c.fill = dark_blue_fill
    for ci in range(1, 12):
        ws2.cell(row=2, column=ci).fill = dark_blue_fill

    # BOQ order + tolerance note
    ws2.merge_cells('A3:K3')
    ws2.cell(row=3, column=1, value='Items listed in BOQ order as received from contractor. 0% tolerance â any quantity mismatch is flagged as DISCREPANCY.').font = Font(name='Arial', size=9, italic=True, color='888888')

    # Column headers per format spec
    headers2 = [
        'Item No.', 'Description', 'Unit', 'BOQ Qty\n(Received)',
        'TraceQ Qty\n(Drawings)', 'Variance\n(+/-)', 'Variance %',
        'Est. Exposure\n(AED)', 'Status', 'Trace ID', 'Notes'
    ]
    row = 4
    for col_idx, h in enumerate(headers2, 1):
        c = ws2.cell(row=row, column=col_idx, value=h)
        c.font = col_header_font
        c.fill = col_header_fill
        c.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
        c.border = thin_border
    ws2.row_dimensions[row].height = 35
    row += 1

    # Data rows â in BOQ order
    item_no = 0
    for idx, comp in enumerate(comparisons):
        item_no += 1
        status = comp['Risk']
        is_alt = (idx % 2 == 1)

        boq_val = comp.get('_boq_qty', 0)
        dwg_val = comp.get('Drawing Qty', 'â')
        diff_val = comp.get('Difference', 'â')
        var_pct = comp.get('Variance %', 'â')
        exp_val = comp.get('_exposure_num', 0) or 0

        vals = [
            item_no,
            comp['Equipment'],
            comp['Unit'],
            boq_val,
            dwg_val,
            diff_val,
            var_pct,
            exp_val if exp_val > 0 else 'â',
            status,
            comp['Trace ID'],
            comp['Notes'],
        ]

        for col_idx, v in enumerate(vals, 1):
            c = ws2.cell(row=row, column=col_idx, value=v)
            c.font = normal_font
            c.border = thin_border

            # Status cell formatting
            if col_idx == 9:
                if status == 'MATCH':
                    c.fill = match_fill
                    c.font = match_status_font
                else:
                    c.fill = discrep_fill
                    c.font = discrep_status_font
            elif is_alt:
                c.fill = alt_fill

            # Trace ID formatting
            if col_idx == 10:
                c.font = trace_font

            # Notes formatting
            if col_idx == 11:
                c.font = notes_font
                c.alignment = Alignment(wrap_text=True, vertical='top')
            elif col_idx in (1, 3, 4, 5, 6, 7, 8, 9, 10):
                c.alignment = Alignment(horizontal='center')
            else:
                c.alignment = Alignment(horizontal='left', wrap_text=True)

            # Number formats
            if col_idx == 4 and isinstance(v, (int, float)):
                c.number_format = '#,##0'
            if col_idx == 8 and isinstance(v, (int, float)):
                c.number_format = '#,##0'
                c.font = aed_font
        row += 1

    # ââ AED Totals Row ââ
    row += 1
    ws2.cell(row=row, column=7, value='TOTAL:').font = bold_font_big
    ws2.cell(row=row, column=7).alignment = Alignment(horizontal='right')
    c = ws2.cell(row=row, column=8, value=comparison_exposure)
    c.font = Font(name='Arial', bold=True, size=12, color=discrep_text)
    c.number_format = '#,##0'
    c.alignment = Alignment(horizontal='center')
    c.border = thin_border
    _apply_border_row(ws2, row, 11)

    # ââ Trace ID Reference Section ââ
    row += 2
    ws2.cell(row=row, column=1, value='TRACE ID REFERENCE').font = Font(name='Arial', bold=True, size=11, color=navy)
    for ci in range(1, 12):
        ws2.cell(row=row, column=ci).fill = section_fill
        ws2.cell(row=row, column=ci).border = thin_border
    row += 1

    ref_headers = ['Trace ID', 'Category', 'Detection Method', 'How to Verify in AutoCAD']
    for col_idx, h in enumerate(ref_headers, 1):
        c = ws2.cell(row=row, column=col_idx, value=h)
        c.font = col_header_font
        c.fill = col_header_fill
        c.alignment = Alignment(horizontal='center')
        c.border = thin_border
    row += 1

    # Add unique trace ID categories
    seen_prefixes = set()
    for comp in comparisons:
        tid = comp['Trace ID']
        prefix = '-'.join(tid.split('-')[:2])
        if prefix in seen_prefixes:
            continue
        seen_prefixes.add(prefix)
        etype = comp.get('_equipment_type', '')
        source_label = comp.get('Detection Source', 'â')
        # Fallback to expected detection method from config when engine didn't detect
        if source_label in ('â', '', None):
            source_label = EXPECTED_DETECTION_METHOD.get(etype, 'â')

        ws2.cell(row=row, column=1, value=f'{prefix}-*').font = trace_font
        ws2.cell(row=row, column=2, value=comp['Equipment']).font = normal_font
        ws2.cell(row=row, column=3, value=source_label).font = normal_font
        ws2.cell(row=row, column=4, value=f'Select layer containing {etype.replace("_", " ")} entities in AutoCAD').font = notes_font
        ws2.cell(row=row, column=4).alignment = Alignment(wrap_text=True)
        _apply_border_row(ws2, row, 4)
        row += 1

    # Column widths
    tab2_widths = [8, 30, 8, 14, 16, 12, 12, 16, 16, 14, 55]
    for i, w in enumerate(tab2_widths):
        ws2.column_dimensions[get_column_letter(i + 1)].width = w

    # âââ TAB 3: MISSING FROM BOQ ââââââââââââââââââââââââââââââââââââââââââââ
    ws3 = wb.create_sheet("Missing from BOQ")
    ws3.sheet_properties.tabColor = 'E65100'

    # Banner
    ws3.merge_cells('A1:H1')
    c = ws3['A1']
    c.value = 'Items Detected in Drawing â Not in BOQ'
    c.font = banner_font
    c.fill = navy_fill
    c.alignment = Alignment(horizontal='center', vertical='center')
    for ci in range(1, 9):
        ws3.cell(row=1, column=ci).fill = navy_fill

    # Context one-liner
    ws3.merge_cells('A2:H2')
    c = ws3['A2']
    c.value = 'The following items were detected on the drawings but do not appear in the BOQ provided. These may represent unpriced scope requiring QS review.'
    c.font = Font(name='Arial', italic=True, size=10, color='666666')
    c.alignment = Alignment(horizontal='left')

    # Column headers per format spec
    headers3 = ['Item', 'Found In Drawing', 'Est. Qty', 'Unit Rate\n(AED)', 'Est. Exposure\n(AED)', 'Status', 'Trace ID', 'Notes']
    row = 4
    for col_idx, h in enumerate(headers3, 1):
        c = ws3.cell(row=row, column=col_idx, value=h)
        c.font = col_header_font
        c.fill = col_header_fill
        c.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
        c.border = thin_border
    ws3.row_dimensions[row].height = 35
    row += 1

    if missing_from_boq:
        for idx, m in enumerate(missing_from_boq):
            qty = m['Drawing Qty']
            unit_rate = m.get('_unit_rate', 0)
            est_exp = m.get('_est_exposure', 0)
            is_alt = (idx % 2 == 1)

            vals = [
                m['Equipment'],
                m['Detection'],
                f"{qty} nos.",
                unit_rate if unit_rate > 0 else 'â',
                est_exp if est_exp > 0 else 'â',
                'MISSING FROM BOQ',
                m['Trace ID'],
                m['Notes'],
            ]

            for col_idx, v in enumerate(vals, 1):
                c = ws3.cell(row=row, column=col_idx, value=v)
                c.font = normal_font
                c.border = thin_border

                if col_idx == 6:
                    c.fill = missing_fill
                    c.font = missing_status_font
                elif is_alt:
                    c.fill = alt_fill

                if col_idx == 7:
                    c.font = trace_font
                if col_idx == 8:
                    c.font = notes_font
                    c.alignment = Alignment(wrap_text=True, vertical='top')
                elif col_idx in (3, 4, 5, 6, 7):
                    c.alignment = Alignment(horizontal='center')
                else:
                    c.alignment = Alignment(horizontal='left')

                if col_idx in (4, 5) and isinstance(v, (int, float)):
                    c.number_format = '#,##0'
                    if col_idx == 5:
                        c.font = aed_font
            row += 1

        # Totals row
        row += 1
        ws3.cell(row=row, column=4, value='MISSING ITEMS TOTAL:').font = bold_font_big
        ws3.cell(row=row, column=4).alignment = Alignment(horizontal='right')
        c = ws3.cell(row=row, column=5, value=missing_exposure)
        c.font = Font(name='Arial', bold=True, size=12, color=discrep_text)
        c.number_format = '#,##0'
        c.alignment = Alignment(horizontal='center')
        c.border = thin_border

        # Maths breakdown (missing items only â combined total lives on Exec Summary)
        row += 1
        ws3.cell(row=row, column=3, value='Missing items total only. See Executive Summary for combined exposure.').font = Font(name='Arial', size=9, italic=True, color='888888')

    else:
        ws3.cell(row=row, column=1, value='No items missing from BOQ.').font = normal_font

    # Disclaimer
    row += 2
    ws3.merge_cells(f'A{row}:H{row}')
    disclaimer = (
        'Unit rates are estimated based on typical UAE HVAC market pricing for indicative purposes only. '
        'Actual costs must be confirmed with project-specific quotations. This report highlights potential '
        'discrepancies for QS review â it does not constitute a formal quantity takeoff or financial advice.'
    )
    c = ws3.cell(row=row, column=1, value=disclaimer)
    c.font = disclaimer_font
    c.alignment = Alignment(wrap_text=True)

    tab3_widths = [22, 22, 12, 14, 16, 20, 14, 50]
    for i, w in enumerate(tab3_widths):
        ws3.column_dimensions[get_column_letter(i + 1)].width = w

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# âââ Page Config ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
st.set_page_config(
    page_title="TraceQ â BOQ Risk Review",
    page_icon="ð",
    layout="wide",
    initial_sidebar_state="expanded",
)

# âââ Branding & Styles âââââââââââââââââââââââââââââââââââââââââââââââââââââââ
st.markdown("""
<style>
    .main-header {
        font-size: 2.2rem;
        font-weight: 700;
        color: #1a1a2e;
        margin-bottom: 0;
    }
    .sub-header {
        font-size: 1.1rem;
        color: #666;
        margin-top: -10px;
        margin-bottom: 30px;
    }
    .metric-card {
        background: #f8f9fa;
        border-radius: 12px;
        padding: 20px;
        border-left: 4px solid #0066cc;
    }
    .risk-high { color: #dc3545; font-weight: 700; }
    .risk-medium { color: #fd7e14; font-weight: 700; }
    .risk-low { color: #28a745; font-weight: 700; }
    .risk-verify { color: #6f42c1; font-weight: 700; }
    .stMetric > div { background: #f8f9fa; border-radius: 10px; padding: 10px; }
</style>
""", unsafe_allow_html=True)

# âââ Header âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
st.markdown('<p class="main-header">ð TraceQ</p>', unsafe_allow_html=True)
st.markdown('<p class="sub-header">BOQ Risk Review Engine â by TechTelligence</p>', unsafe_allow_html=True)

# âââ Sidebar â Page Navigation âââââââââââââââââââââââââââââââââââââââââââââââ
with st.sidebar:
    st.markdown("### Navigation")
    page = st.radio(
        "Select page:",
        ["Engine Analysis", "Upload Validator Response", "Generate Client Report"],
        label_visibility="collapsed",
    )
    st.markdown("---")


# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
# VALIDATOR RESPONSE PAGE
# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ

def parse_validator_xlsx(file_bytes, filename):
    """Parse a validator submission XLSX and return structured data from 3 tabs."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet_names_upper = {s.upper(): s for s in wb.sheetnames}

    result = {
        'filename': filename,
        'boq_comparison': [],
        'missing_from_boq': [],
        'discoveries': [],
    }

    # ââ BOQ Comparison tab ââ
    boq_tab = None
    for key in ['BOQ COMPARISON', 'BOQ_COMPARISON']:
        if key in sheet_names_upper:
            boq_tab = wb[sheet_names_upper[key]]
            break
    if boq_tab is None:
        for sn in wb.sheetnames:
            if 'BOQ' in sn.upper() and 'COMPARISON' in sn.upper():
                boq_tab = wb[sn]
                break
    if boq_tab is None and len(wb.sheetnames) >= 2:
        boq_tab = wb[wb.sheetnames[1]]  # fallback: second sheet

    if boq_tab:
        # Find header row (look for "Equipment" or "#" in first 5 rows)
        header_row = 2  # default
        _found_header = False
        for r in range(1, 6):
            if _found_header:
                break
            for c in range(1, 10):
                val = boq_tab.cell(row=r, column=c).value
                if val and 'EQUIPMENT' in str(val).upper():
                    header_row = r
                    _found_header = True
                    break

        # Read data rows starting after header
        for r in range(header_row + 1, boq_tab.max_row + 1):
            row_num = boq_tab.cell(row=r, column=1).value
            if row_num is None:
                continue
            try:
                int(row_num)
            except (ValueError, TypeError):
                continue

            equipment = str(boq_tab.cell(row=r, column=2).value or '').strip()
            if not equipment:
                continue

            unit = str(boq_tab.cell(row=r, column=3).value or '').strip()
            boq_qty = boq_tab.cell(row=r, column=4).value
            engine_count = boq_tab.cell(row=r, column=5).value
            variance = boq_tab.cell(row=r, column=6).value
            agree = str(boq_tab.cell(row=r, column=7).value or '').strip()
            comments = str(boq_tab.cell(row=r, column=8).value or '').strip()

            # Try to get numeric values
            try:
                boq_qty = float(boq_qty) if boq_qty is not None else None
            except (ValueError, TypeError):
                pass
            try:
                engine_count = float(engine_count) if engine_count is not None else None
            except (ValueError, TypeError):
                pass

            result['boq_comparison'].append({
                'equipment': equipment,
                'unit': unit,
                'boq_qty': boq_qty,
                'engine_count': engine_count,
                'variance': variance,
                'agree': agree.upper() if agree else '',
                'comments': comments,
            })

    # ââ Missing from BOQ tab ââ
    missing_tab = None
    for sn in wb.sheetnames:
        if 'MISSING' in sn.upper():
            missing_tab = wb[sn]
            break
    if missing_tab is None and len(wb.sheetnames) >= 3:
        missing_tab = wb[wb.sheetnames[2]]

    if missing_tab:
        header_row = 2
        _found_header = False
        for r in range(1, 6):
            if _found_header:
                break
            for c in range(1, 10):
                val = missing_tab.cell(row=r, column=c).value
                if val and 'EQUIPMENT' in str(val).upper():
                    header_row = r
                    _found_header = True
                    break

        for r in range(header_row + 1, missing_tab.max_row + 1):
            row_num = missing_tab.cell(row=r, column=1).value
            if row_num is None:
                continue
            try:
                int(row_num)
            except (ValueError, TypeError):
                continue

            equipment = str(missing_tab.cell(row=r, column=2).value or '').strip()
            if not equipment:
                continue

            unit = str(missing_tab.cell(row=r, column=3).value or '').strip()
            engine_count = missing_tab.cell(row=r, column=4).value
            detection_source = str(missing_tab.cell(row=r, column=5).value or '').strip()
            agree = str(missing_tab.cell(row=r, column=6).value or '').strip()
            comments = str(missing_tab.cell(row=r, column=7).value or '').strip()

            result['missing_from_boq'].append({
                'equipment': equipment,
                'unit': unit,
                'engine_count': engine_count,
                'detection_source': detection_source,
                'agree': agree.upper() if agree else '',
                'comments': comments,
            })

    # ââ Validator Discoveries tab ââ
    disc_tab = None
    for sn in wb.sheetnames:
        if 'DISCOVER' in sn.upper():
            disc_tab = wb[sn]
            break
    if disc_tab is None and len(wb.sheetnames) >= 4:
        disc_tab = wb[wb.sheetnames[3]]

    if disc_tab:
        header_row = 3  # Discoveries typically has instructions in rows 1-2
        _found_header = False
        for r in range(1, 6):
            if _found_header:
                break
            for c in range(1, 8):
                val = disc_tab.cell(row=r, column=c).value
                if val and 'EQUIPMENT' in str(val).upper():
                    header_row = r
                    _found_header = True
                    break

        for r in range(header_row + 1, disc_tab.max_row + 1):
            row_num = disc_tab.cell(row=r, column=1).value
            if row_num is None:
                continue
            equipment = str(disc_tab.cell(row=r, column=2).value or '').strip()
            if not equipment:
                continue

            unit = str(disc_tab.cell(row=r, column=3).value or '').strip()
            count = disc_tab.cell(row=r, column=4).value
            location = str(disc_tab.cell(row=r, column=5).value or '').strip()
            reason = str(disc_tab.cell(row=r, column=6).value or '').strip()

            result['discoveries'].append({
                'equipment': equipment,
                'unit': unit,
                'count': count,
                'location': location,
                'reason_missed': reason,
            })

    wb.close()
    return result


def compare_validators(v1, v2):
    """Compare two validator submissions. Returns agreements, divergences, engine error candidates."""
    agreements = []
    divergences = []
    engine_errors = []

    # ââ Compare BOQ Comparison verdicts ââ
    # Build lookup by equipment name (normalised)
    v1_boq = {item['equipment'].upper(): item for item in v1['boq_comparison']}
    v2_boq = {item['equipment'].upper(): item for item in v2['boq_comparison']}

    all_equipment = set(v1_boq.keys()) | set(v2_boq.keys())
    for equip in sorted(all_equipment):
        item1 = v1_boq.get(equip, {})
        item2 = v2_boq.get(equip, {})

        agree1 = item1.get('agree', '').upper()
        agree2 = item2.get('agree', '').upper()

        # Normalise Yes/No
        def _norm(val):
            if not val:
                return 'MISSING'
            if val in ('YES', 'Y'):
                return 'YES'
            if val in ('NO', 'N'):
                return 'NO'
            if val.startswith('PARTIAL'):
                return 'PARTIAL'
            return val

        n1, n2 = _norm(agree1), _norm(agree2)

        row = {
            'equipment': item1.get('equipment', item2.get('equipment', equip)),
            'boq_qty': item1.get('boq_qty') or item2.get('boq_qty'),
            'engine_count': item1.get('engine_count') or item2.get('engine_count'),
            'v1_agree': n1,
            'v2_agree': n2,
            'v1_comments': item1.get('comments', ''),
            'v2_comments': item2.get('comments', ''),
        }

        if n1 == n2:
            agreements.append(row)
            # Both say NO = engine error candidate
            if n1 == 'NO':
                engine_errors.append({
                    'equipment': row['equipment'],
                    'issue': 'Both validators disagree with engine count',
                    'v1_comments': row['v1_comments'],
                    'v2_comments': row['v2_comments'],
                    'boq_qty': row['boq_qty'],
                    'engine_count': row['engine_count'],
                })
        else:
            divergences.append(row)

    # ââ Compare Missing from BOQ verdicts ââ
    v1_miss = {item['equipment'].upper(): item for item in v1['missing_from_boq']}
    v2_miss = {item['equipment'].upper(): item for item in v2['missing_from_boq']}

    all_missing = set(v1_miss.keys()) | set(v2_miss.keys())
    for equip in sorted(all_missing):
        item1 = v1_miss.get(equip, {})
        item2 = v2_miss.get(equip, {})

        agree1 = item1.get('agree', '').upper()
        agree2 = item2.get('agree', '').upper()

        def _norm(val):
            if not val:
                return 'MISSING'
            if val in ('YES', 'Y'):
                return 'YES'
            if val in ('NO', 'N'):
                return 'NO'
            return val

        n1, n2 = _norm(agree1), _norm(agree2)

        row = {
            'equipment': item1.get('equipment', item2.get('equipment', equip)),
            'engine_count': item1.get('engine_count') or item2.get('engine_count'),
            'v1_agree': n1,
            'v2_agree': n2,
            'v1_comments': item1.get('comments', ''),
            'v2_comments': item2.get('comments', ''),
            'section': 'Missing from BOQ',
        }

        if n1 == n2:
            agreements.append(row)
            if n1 == 'NO':
                engine_errors.append({
                    'equipment': row['equipment'],
                    'issue': 'Both validators say item is NOT missing from BOQ (false positive)',
                    'v1_comments': row['v1_comments'],
                    'v2_comments': row['v2_comments'],
                    'engine_count': row['engine_count'],
                })
        else:
            divergences.append(row)

    # ââ Collect discoveries from both ââ
    all_discoveries = []
    for d in v1.get('discoveries', []):
        d['source_validator'] = v1['filename']
        all_discoveries.append(d)
    for d in v2.get('discoveries', []):
        d['source_validator'] = v2['filename']
        all_discoveries.append(d)

    return {
        'agreements': agreements,
        'divergences': divergences,
        'engine_errors': engine_errors,
        'discoveries': all_discoveries,
    }


def write_to_l1_tracker(tracker_bytes, job_id, validator_data_list, comparison=None):
    """
    Write validator data to L1 Feedback Tracker XLSX.
    Populates QS Feedback tab and Engine Improvement Log tab.
    Returns updated workbook bytes.
    """
    wb = openpyxl.load_workbook(io.BytesIO(tracker_bytes))

    # ââ QS Feedback tab ââ
    qs_tab = None
    for sn in wb.sheetnames:
        if 'QS' in sn.upper() and 'FEEDBACK' in sn.upper():
            qs_tab = wb[sn]
            break

    if qs_tab:
        # Find next empty row (headers at row 4, data starts row 5)
        next_row = 5
        while qs_tab.cell(row=next_row, column=1).value is not None:
            next_row += 1

        item_no = 1
        for vdata in validator_data_list:
            validator_name = vdata['filename'].replace('.xlsx', '').replace('.XLSX', '')

            # Write BOQ Comparison items
            for item in vdata['boq_comparison']:
                qs_tab.cell(row=next_row, column=1, value=job_id)  # Job ID
                qs_tab.cell(row=next_row, column=2, value=item_no)  # Item No.
                qs_tab.cell(row=next_row, column=3, value=item['equipment'])  # Equipment Type
                qs_tab.cell(row=next_row, column=4, value=item.get('engine_count'))  # Engine Count
                qs_tab.cell(row=next_row, column=5, value=item.get('boq_qty'))  # BOQ Count
                qs_tab.cell(row=next_row, column=6, value=item.get('agree', ''))  # Engine Status mapped to QS Agrees?
                qs_tab.cell(row=next_row, column=7, value=item.get('agree', ''))  # QS Agrees?
                qs_tab.cell(row=next_row, column=9, value=item.get('comments', ''))  # QS Reasoning
                # Is This a False Positive? â YES if validator says NO
                is_fp = 'YES' if item.get('agree', '').upper() == 'NO' else 'NO'
                qs_tab.cell(row=next_row, column=10, value=is_fp)
                qs_tab.cell(row=next_row, column=14, value=validator_name)  # Loop 2 Note (validator ID)
                next_row += 1
                item_no += 1

            # Write Missing from BOQ items
            for item in vdata['missing_from_boq']:
                qs_tab.cell(row=next_row, column=1, value=job_id)
                qs_tab.cell(row=next_row, column=2, value=item_no)
                qs_tab.cell(row=next_row, column=3, value=item['equipment'])
                qs_tab.cell(row=next_row, column=4, value=item.get('engine_count'))
                qs_tab.cell(row=next_row, column=6, value='MISSING')
                qs_tab.cell(row=next_row, column=7, value=item.get('agree', ''))
                qs_tab.cell(row=next_row, column=9, value=item.get('comments', ''))
                is_fp = 'YES' if item.get('agree', '').upper() == 'NO' else 'NO'
                qs_tab.cell(row=next_row, column=10, value=is_fp)
                qs_tab.cell(row=next_row, column=14, value=validator_name)
                next_row += 1
                item_no += 1

    # ââ Engine Improvement Log tab ââ
    eng_tab = None
    for sn in wb.sheetnames:
        if 'ENGINE' in sn.upper() and 'IMPROVEMENT' in sn.upper():
            eng_tab = wb[sn]
            break
    if eng_tab is None:
        for sn in wb.sheetnames:
            if 'ENGINE' in sn.upper():
                eng_tab = wb[sn]
                break

    if eng_tab and comparison and comparison.get('engine_errors'):
        next_row = 5
        while eng_tab.cell(row=next_row, column=1).value is not None:
            next_row += 1

        fix_id = 1
        for err in comparison['engine_errors']:
            eng_tab.cell(row=next_row, column=1, value=f"FIX-{job_id}-{fix_id:03d}")  # Fix ID
            eng_tab.cell(row=next_row, column=2, value=datetime.now().strftime('%Y-%m-%d'))  # Date Logged
            eng_tab.cell(row=next_row, column=3, value=err['equipment'])  # Equipment Type
            eng_tab.cell(row=next_row, column=4, value=err['issue'])  # Issue Description
            eng_tab.cell(row=next_row, column=6, value=job_id)  # Source Job IDs
            eng_tab.cell(row=next_row, column=7, value='OPEN')  # Fix Status
            next_row += 1
            fix_id += 1

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    wb.close()
    return output.getvalue()


def parse_engine_report(file_bytes):
    """
    Parse the original engine-generated BOQ Report XLSX.
    Recovers comparisons, missing_from_boq, drawing name, BOQ name, and metadata.
    Returns dict with all recovered data.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    result = {
        'drawing_name': '',
        'boq_name': '',
        'comparisons': [],
        'missing_from_boq': [],
    }

    # ââ Parse Executive Summary for metadata ââ
    ws1 = wb.worksheets[0] if wb.worksheets else None
    if ws1:
        for r in range(4, 10):
            label = str(ws1.cell(row=r, column=1).value or '').strip()
            val = str(ws1.cell(row=r, column=2).value or '').strip()
            if 'Drawing' in label:
                result['drawing_name'] = val
            elif 'BOQ' in label:
                result['boq_name'] = val

    # ââ Parse BOQ Comparison tab ââ
    ws2 = None
    for sn in wb.sheetnames:
        if 'BOQ' in sn.upper() and 'COMPARISON' in sn.upper():
            ws2 = wb[sn]
            break
    if ws2 is None and len(wb.sheetnames) >= 2:
        ws2 = wb.worksheets[1]

    if ws2:
        # Find header row (look for "Description" or "Item No.")
        header_row = 4  # default per generate_excel_report
        for r in range(1, 8):
            for c in range(1, 12):
                val = ws2.cell(row=r, column=c).value
                if val and 'DESCRIPTION' in str(val).upper():
                    header_row = r
                    break

        for r in range(header_row + 1, ws2.max_row + 1):
            desc = ws2.cell(row=r, column=2).value
            if desc is None or str(desc).strip() == '':
                # Check if this is the totals row or end
                col7 = ws2.cell(row=r, column=7).value
                if col7 and 'TOTAL' in str(col7).upper():
                    break
                continue

            desc_str = str(desc).strip()
            unit = str(ws2.cell(row=r, column=3).value or '').strip()
            boq_qty = ws2.cell(row=r, column=4).value
            drawing_qty = ws2.cell(row=r, column=5).value
            variance = ws2.cell(row=r, column=6).value
            var_pct = ws2.cell(row=r, column=7).value
            exposure = ws2.cell(row=r, column=8).value
            status = str(ws2.cell(row=r, column=9).value or '').strip()
            trace_id = str(ws2.cell(row=r, column=10).value or '').strip()
            notes = str(ws2.cell(row=r, column=11).value or '').strip()

            # Recover numeric values
            try:
                boq_qty = float(boq_qty) if boq_qty is not None else 0
            except (ValueError, TypeError):
                boq_qty = 0
            try:
                drawing_qty = float(drawing_qty) if drawing_qty is not None else 0
            except (ValueError, TypeError):
                drawing_qty = 0
            try:
                exposure_num = float(exposure) if exposure not in (None, '', 'â') else 0
            except (ValueError, TypeError):
                exposure_num = 0

            # Calculate unit rate from exposure and variance
            diff = abs(drawing_qty - boq_qty) if drawing_qty and boq_qty else 0
            unit_rate = exposure_num / diff if diff > 0 and exposure_num > 0 else 0

            # Classify equipment type
            etype, label = _classify_description(desc_str)

            result['comparisons'].append({
                'Equipment': desc_str,
                'Unit': unit,
                'BOQ Qty': boq_qty,
                'Drawing Qty': drawing_qty,
                'Difference': variance,
                'Variance %': var_pct,
                'Risk': status,
                'Trace ID': trace_id,
                'Notes': notes,
                'Exposure (AED)': f'AED {exposure_num:,.0f}' if exposure_num > 0 else 'â',
                '_exposure_num': exposure_num,
                '_boq_qty': boq_qty,
                '_unit_rate': unit_rate,
                '_equipment_type': etype or '',
            })

    # ââ Parse Missing from BOQ tab ââ
    ws3 = None
    for sn in wb.sheetnames:
        if 'MISSING' in sn.upper():
            ws3 = wb[sn]
            break
    if ws3 is None and len(wb.sheetnames) >= 3:
        ws3 = wb.worksheets[2]

    if ws3:
        header_row = 4
        for r in range(1, 8):
            for c in range(1, 9):
                val = ws3.cell(row=r, column=c).value
                if val and 'ITEM' in str(val).upper():
                    header_row = r
                    break

        for r in range(header_row + 1, ws3.max_row + 1):
            equip = ws3.cell(row=r, column=1).value
            if equip is None or str(equip).strip() == '':
                continue

            equip_str = str(equip).strip()
            if 'TOTAL' in equip_str.upper() or 'DISCLAIMER' in equip_str.upper():
                break
            if 'Unit rates' in equip_str or 'Actual costs' in equip_str:
                break

            detection = str(ws3.cell(row=r, column=2).value or '').strip()
            qty_str = str(ws3.cell(row=r, column=3).value or '').strip()
            unit_rate = ws3.cell(row=r, column=4).value
            est_exp = ws3.cell(row=r, column=5).value
            status = str(ws3.cell(row=r, column=6).value or '').strip()
            trace_id = str(ws3.cell(row=r, column=7).value or '').strip()
            notes = str(ws3.cell(row=r, column=8).value or '').strip()

            # Extract numeric qty from "X nos." format
            import re as _re
            qty_match = _re.search(r'([\d,.]+)', qty_str)
            qty = float(qty_match.group(1).replace(',', '')) if qty_match else 0

            try:
                unit_rate_num = float(unit_rate) if unit_rate not in (None, '', 'â') else 0
            except (ValueError, TypeError):
                unit_rate_num = 0
            try:
                est_exp_num = float(est_exp) if est_exp not in (None, '', 'â') else 0
            except (ValueError, TypeError):
                est_exp_num = 0

            etype, label = _classify_description(equip_str)

            result['missing_from_boq'].append({
                'Equipment': equip_str,
                'Detection': detection,
                'Drawing Qty': qty,
                'Trace ID': trace_id,
                'Notes': notes,
                '_unit_rate': unit_rate_num,
                '_est_exposure': est_exp_num,
                '_equipment_type': etype or '',
            })

    wb.close()
    return result


def extract_corrected_count(comments):
    """Extract a corrected count from validator comments if present."""
    import re as _re
    if not comments:
        return None
    # Patterns like "correct count: 15", "should be 15", "actual: 15", "actual count is 32"
    patterns = [
        r'correct(?:ed)?\s*(?:count|qty|quantity)?\s*[:\-=]\s*(\d+)',
        r'should\s+be\s+(\d+)',
        r'actual\s*(?:count|qty|quantity)?\s*(?:is|[:\-=])\s*(\d+)',
        r'count\s*(?:is|[:\-=])\s*(\d+)',
    ]
    for pat in patterns:
        m = _re.search(pat, comments, _re.IGNORECASE)
        if m:
            return float(m.group(1))
    return None


def match_equipment_name(name1, name2):
    """Fuzzy match two equipment names. Returns True if they likely refer to the same item."""
    import re as _re
    def _normalize(s):
        s = s.upper().strip()
        s = _re.sub(r'[^A-Z0-9\s]', '', s)
        s = _re.sub(r'\s+', ' ', s)
        return s

    n1 = _normalize(name1)
    n2 = _normalize(name2)

    if n1 == n2:
        return True
    if n1 in n2 or n2 in n1:
        return True
    # Check if one is an abbreviation of the other (first letters match)
    words1 = n1.split()
    words2 = n2.split()
    if len(words1) == 1 and len(words2) > 1:
        abbrev = ''.join(w[0] for w in words2)
        if n1 == abbrev:
            return True
    if len(words2) == 1 and len(words1) > 1:
        abbrev = ''.join(w[0] for w in words1)
        if n2 == abbrev:
            return True
    return False


def merge_validated_data(engine_data, validator_data_list, comparison_result=None):
    """
    Merge engine report data with validator verdicts to produce client report data.
    Returns (comparisons, missing_from_boq, metadata) in generate_excel_report format.

    Rules:
    - Validator YES: keep engine item
    - Validator NO (both): exclude (engine error)
    - Validators disagree: INCLUDE (conservative rule)
    - Validator-corrected counts: use corrected count
    - Discoveries: fold into missing_from_boq
    """
    merged_comparisons = []
    merged_missing = []
    metadata = {
        'validation_method': 'Dual Validator (consensus)' if len(validator_data_list) == 2 else 'Single Validator',
        'validator_names': [v['filename'].replace('.xlsx', '').replace('.XLSX', '') for v in validator_data_list],
        'engine_errors_excluded': 0,
        'validator_corrections': 0,
        'discoveries_added': 0,
        'conservative_includes': 0,
    }

    is_dual = len(validator_data_list) == 2 and comparison_result is not None

    # ââ Build validator verdict lookups ââ
    if is_dual:
        # Use comparison_result for verdicts
        boq_agreements = {item['equipment'].upper(): item for item in comparison_result.get('agreements', [])}
        boq_divergences = {item['equipment'].upper(): item for item in comparison_result.get('divergences', [])}
        boq_engine_errors = {item['equipment'].upper(): item for item in comparison_result.get('engine_errors', [])}

        # Missing from BOQ â separate from BOQ comparison
        miss_agreements = {}
        miss_divergences = {}
        miss_engine_errors = {}
        for item in comparison_result.get('agreements', []):
            if item.get('section') == 'Missing from BOQ':
                miss_agreements[item['equipment'].upper()] = item
        for item in comparison_result.get('divergences', []):
            if item.get('section') == 'Missing from BOQ':
                miss_divergences[item['equipment'].upper()] = item
        for item in comparison_result.get('engine_errors', []):
            if 'false positive' in item.get('issue', '').lower():
                miss_engine_errors[item['equipment'].upper()] = item
    else:
        # Single validator â use raw parsed data
        v1 = validator_data_list[0]
        v1_boq = {item['equipment'].upper(): item for item in v1['boq_comparison']}
        v1_miss = {item['equipment'].upper(): item for item in v1['missing_from_boq']}

    # ââ Process BOQ Comparison items ââ
    for comp in engine_data['comparisons']:
        equip_upper = comp['Equipment'].upper()

        if is_dual:
            # Check engine errors first (both said NO)
            if equip_upper in boq_engine_errors:
                metadata['engine_errors_excluded'] += 1
                continue  # Exclude from client report

            # Check divergences (conservative include)
            if equip_upper in boq_divergences:
                div = boq_divergences[equip_upper]
                comp['Notes'] = (comp['Notes'] + ' | Validator disagreement â included for review.').strip(' | ')
                comp['_confidence'] = 'Conservative Include'
                metadata['conservative_includes'] += 1
                merged_comparisons.append(comp)
                continue

            # Check agreements
            if equip_upper in boq_agreements:
                agr = boq_agreements[equip_upper]
                # Both said YES â confirmed
                if agr.get('v1_agree') == 'YES' and agr.get('v2_agree') == 'YES':
                    comp['_confidence'] = 'Validator Confirmed'
                # Both said NO (already caught as engine error above, but safety)
                elif agr.get('v1_agree') == 'NO' and agr.get('v2_agree') == 'NO':
                    metadata['engine_errors_excluded'] += 1
                    continue
                else:
                    # Both PARTIAL or other
                    comments = '; '.join(filter(None, [agr.get('v1_comments', ''), agr.get('v2_comments', '')]))
                    corrected = extract_corrected_count(comments)
                    if corrected is not None:
                        old_qty = comp.get('Drawing Qty', 0) or 0
                        comp['Drawing Qty'] = corrected
                        comp['Difference'] = corrected - comp['_boq_qty']
                        if comp['_boq_qty'] > 0:
                            comp['Variance %'] = f"{abs(comp['Difference']) / comp['_boq_qty'] * 100:.0f}%"
                        comp['_exposure_num'] = abs(comp['Difference']) * comp.get('_unit_rate', 0)
                        comp['Notes'] = (comp['Notes'] + f' | Validator corrected: {int(old_qty)} â {int(corrected)}').strip(' | ')
                        comp['_confidence'] = 'Validator Corrected'
                        metadata['validator_corrections'] += 1
                    else:
                        if comments:
                            comp['Notes'] = (comp['Notes'] + f' | Validator: {comments[:100]}').strip(' | ')
                        comp['_confidence'] = 'Validator Confirmed'
                merged_comparisons.append(comp)
                continue

            # No validator match found â include as-is
            comp['_confidence'] = 'Engine Only'
            merged_comparisons.append(comp)

        else:
            # Single validator path
            v_item = v1_boq.get(equip_upper)
            if v_item is None:
                # Try fuzzy match
                for k, v in v1_boq.items():
                    if match_equipment_name(equip_upper, k):
                        v_item = v
                        break

            if v_item is None:
                comp['_confidence'] = 'Engine Only'
                merged_comparisons.append(comp)
                continue

            verdict = v_item.get('agree', '').upper()
            comments = v_item.get('comments', '')

            if verdict == 'NO':
                # Check for corrected count
                corrected = extract_corrected_count(comments)
                if corrected is not None:
                    old_qty = comp.get('Drawing Qty', 0) or 0
                    comp['Drawing Qty'] = corrected
                    comp['Difference'] = corrected - comp['_boq_qty']
                    if comp['_boq_qty'] > 0:
                        comp['Variance %'] = f"{abs(comp['Difference']) / comp['_boq_qty'] * 100:.0f}%"
                    comp['_exposure_num'] = abs(comp['Difference']) * comp.get('_unit_rate', 0)
                    comp['Notes'] = (comp['Notes'] + f' | Validator corrected: {int(old_qty)} â {int(corrected)}').strip(' | ')
                    comp['_confidence'] = 'Validator Corrected'
                    metadata['validator_corrections'] += 1
                    merged_comparisons.append(comp)
                else:
                    # Exclude â engine error
                    metadata['engine_errors_excluded'] += 1
                continue
            elif verdict == 'YES':
                comp['_confidence'] = 'Validator Confirmed'
                if comments:
                    comp['Notes'] = (comp['Notes'] + f' | Validator: {comments[:100]}').strip(' | ')
                merged_comparisons.append(comp)
            else:
                # PARTIAL or other â include with note
                comp['_confidence'] = 'Validator Confirmed'
                if comments:
                    comp['Notes'] = (comp['Notes'] + f' | Validator: {comments[:100]}').strip(' | ')
                merged_comparisons.append(comp)

    # ââ Process Missing from BOQ items ââ
    for m in engine_data['missing_from_boq']:
        equip_upper = m['Equipment'].upper()

        if is_dual:
            if equip_upper in miss_engine_errors:
                metadata['engine_errors_excluded'] += 1
                continue

            if equip_upper in miss_divergences:
                m['Notes'] = (m['Notes'] + ' | Validator disagreement â included for review.').strip(' | ')
                metadata['conservative_includes'] += 1
                merged_missing.append(m)
                continue

            if equip_upper in miss_agreements:
                agr = miss_agreements[equip_upper]
                if agr.get('v1_agree') == 'NO' and agr.get('v2_agree') == 'NO':
                    metadata['engine_errors_excluded'] += 1
                    continue
                merged_missing.append(m)
                continue

            # No verdict found â include as-is
            merged_missing.append(m)
        else:
            v_item = v1_miss.get(equip_upper)
            if v_item is None:
                for k, v in v1_miss.items():
                    if match_equipment_name(equip_upper, k):
                        v_item = v
                        break

            if v_item is None:
                merged_missing.append(m)
                continue

            verdict = v_item.get('agree', '').upper()
            if verdict == 'NO':
                metadata['engine_errors_excluded'] += 1
                continue
            else:
                if v_item.get('comments'):
                    m['Notes'] = (m['Notes'] + f' | Validator: {v_item["comments"][:100]}').strip(' | ')
                merged_missing.append(m)

    # ââ Fold discoveries into Missing from BOQ ââ
    discoveries = []
    if is_dual and comparison_result:
        discoveries = comparison_result.get('discoveries', [])
    elif not is_dual and validator_data_list:
        discoveries = validator_data_list[0].get('discoveries', [])

    # Find the highest trace ID number to continue the sequence
    max_trace_num = 0
    import re as _re
    for comp in merged_comparisons + merged_missing:
        tid = comp.get('Trace ID', '')
        nums = _re.findall(r'(\d+)$', tid)
        if nums:
            max_trace_num = max(max_trace_num, int(nums[0]))

    # Get trace ID prefix from existing items
    trace_prefix = 'TQ'
    for comp in engine_data['comparisons']:
        tid = comp.get('Trace ID', '')
        parts = tid.split('-')
        if len(parts) >= 2:
            trace_prefix = '-'.join(parts[:-1])
            break

    for disc in discoveries:
        max_trace_num += 1
        equip = disc.get('equipment', '')
        qty = disc.get('count', 0)
        try:
            qty = float(qty) if qty is not None else 0
        except (ValueError, TypeError):
            qty = 0

        # Try to find a unit rate
        etype, label = _classify_description(equip)
        unit_rate = UAE_UNIT_RATES.get(etype, 0) if etype else 0
        est_exp = qty * unit_rate

        location = disc.get('location', '')
        reason = disc.get('reason_missed', '')
        note_parts = []
        if location:
            note_parts.append(location)
        if reason:
            note_parts.append(reason)

        merged_missing.append({
            'Equipment': equip,
            'Detection': 'Validator Discovery',
            'Drawing Qty': qty,
            'Trace ID': f'{trace_prefix}-{max_trace_num:03d}',
            'Notes': ' | '.join(note_parts) if note_parts else 'Identified by validator during manual review.',
            '_unit_rate': unit_rate,
            '_est_exposure': est_exp,
            '_equipment_type': etype or '',
        })
        metadata['discoveries_added'] += 1

    return merged_comparisons, merged_missing, metadata


def render_client_report_page():
    """Render the Generate Client Report page."""

    with st.sidebar:
        st.markdown("### Client Report")
        st.markdown("Generate the final client-facing report from engine output + validator responses.")

        engine_report_file = st.file_uploader(
            "Original Engine Report (XLSX)",
            type=["xlsx"],
            help="Upload the TraceQ BOQ Report XLSX downloaded from the Engine Analysis page.",
            key="engine_report_upload",
        )

        validator_files_cr = st.file_uploader(
            "Validator Submission(s)",
            type=["xlsx"],
            accept_multiple_files=True,
            help="Upload the completed XLSX file(s) returned by your validators.",
            key="validator_upload_cr",
        )

        job_id_cr = st.text_input("Job ID", value="TQ-JOB-001", help="Job identifier for the report.", key="job_id_cr")
        client_name = st.text_input("Client Name (optional)", value="", help="Client name for report header.", key="client_name_cr")

    # ââ Main content ââ
    st.markdown("---")

    # Guide
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown("### 1. Upload engine report")
        st.markdown("Drag in the `TraceQ_BOQ_Report` XLSX from the Engine Analysis page.")
    with col2:
        st.markdown("### 2. Upload validator response(s)")
        st.markdown("Drag in 1 or 2 completed validator XLSX files from the sidebar.")
    with col3:
        st.markdown("### 3. Download client report")
        st.markdown("Review the preview, then generate and download the polished report.")

    st.markdown("---")

    if not engine_report_file:
        st.info("Upload the original engine report XLSX in the sidebar to get started.")
        return

    if not validator_files_cr or len(validator_files_cr) == 0:
        st.info("Upload at least one validator submission in the sidebar.")
        return

    if len(validator_files_cr) > 2:
        st.warning("Maximum 2 validator files. Only the first 2 will be used.")
        validator_files_cr = validator_files_cr[:2]

    # ââ Parse engine report ââ
    with st.spinner("Parsing engine report..."):
        engine_bytes = engine_report_file.read()
        engine_data = parse_engine_report(engine_bytes)

    if not engine_data['comparisons'] and not engine_data['missing_from_boq']:
        st.error("Could not parse the engine report. Please ensure this is a TraceQ BOQ Report XLSX from the Engine Analysis page.")
        return

    # ââ Parse validator submissions ââ
    parsed_validators = []
    for vf in validator_files_cr:
        with st.spinner(f"Parsing {vf.name}..."):
            vbytes = vf.read()
            parsed = parse_validator_xlsx(vbytes, vf.name)
            parsed_validators.append(parsed)

    # ââ Run comparison if dual ââ
    comparison_result = None
    if len(parsed_validators) == 2:
        comparison_result = compare_validators(parsed_validators[0], parsed_validators[1])

    # ââ Merge ââ
    with st.spinner("Merging validated data..."):
        merged_comparisons, merged_missing, metadata = merge_validated_data(
            engine_data, parsed_validators, comparison_result
        )

    # ââ Preview stats ââ
    st.markdown("### Merge preview")

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("BOQ items (final)", len(merged_comparisons))
    c2.metric("Engine errors excluded", metadata['engine_errors_excluded'])
    c3.metric("Validator corrections", metadata['validator_corrections'])
    c4.metric("Discoveries added", metadata['discoveries_added'])

    if metadata['conservative_includes'] > 0:
        st.caption(f"Conservative includes (validator disagreement): {metadata['conservative_includes']}")

    st.caption(f"Validation: {metadata['validation_method']} â {', '.join(metadata['validator_names'])}")

    # ââ Expandable previews ââ
    with st.expander(f"BOQ comparison items ({len(merged_comparisons)} items)", expanded=False):
        if merged_comparisons:
            import pandas as pd
            df = pd.DataFrame([{
                'Equipment': c['Equipment'],
                'BOQ Qty': c.get('_boq_qty', ''),
                'Verified Qty': c.get('Drawing Qty', ''),
                'Status': c.get('Risk', ''),
                'Confidence': c.get('_confidence', ''),
            } for c in merged_comparisons])
            st.dataframe(df, use_container_width=True, hide_index=True)

    with st.expander(f"Missing from BOQ ({len(merged_missing)} items)", expanded=False):
        if merged_missing:
            import pandas as pd
            df = pd.DataFrame([{
                'Equipment': m['Equipment'],
                'Qty': m.get('Drawing Qty', ''),
                'Source': m.get('Detection', ''),
                'Trace ID': m.get('Trace ID', ''),
            } for m in merged_missing])
            st.dataframe(df, use_container_width=True, hide_index=True)

    excluded_count = metadata['engine_errors_excluded']
    if excluded_count > 0:
        with st.expander(f"Excluded items â engine errors ({excluded_count})", expanded=False):
            st.markdown("These items were removed because validator(s) confirmed the engine count was incorrect.")

    # ââ Generate button ââ
    st.markdown("---")
    if st.button("Generate Client Report", type="primary"):
        with st.spinner("Generating polished client report..."):
            # Add validation metadata to comparisons for report
            drawing_name = engine_data.get('drawing_name', 'Unknown')
            boq_name = engine_data.get('boq_name', 'Unknown')

            # Call the existing report generator with merged data
            report_bytes = generate_excel_report(
                comparisons=merged_comparisons,
                missing_from_boq=merged_missing,
                boq_items=[],  # Not needed for the report output
                drawing_name=drawing_name,
                boq_name=boq_name,
                validation_metadata=metadata,
            )

            st.success(f"Client report generated! {len(merged_comparisons)} BOQ items, {len(merged_missing)} missing items.")

            safe_job_id = job_id_cr.replace(' ', '_') if job_id_cr else 'report'
            st.download_button(
                label="Download Client Report",
                data=report_bytes,
                file_name=f"TraceQ_Client_Report_{safe_job_id}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.xml",
            )


def render_validator_page():
    """Render the Upload Validator Response page."""

    with st.sidebar:
        st.markdown("### Validator Response")
        st.markdown("Upload 1 or 2 completed validator XLSX files to review and log feedback.")

        validator_files = st.file_uploader(
            "Validator XLSX Submission(s)",
            type=["xlsx"],
            accept_multiple_files=True,
            help="Upload the completed XLSX file(s) returned by your validators.",
            key="validator_upload",
        )

        job_id = st.text_input("Job ID", value="TQ-TRAIN-S5", help="Job identifier for L1 Tracker logging.")

        tracker_file = st.file_uploader(
            "L1 Feedback Tracker (override)",
            type=["xlsx"],
            help="Optional: upload a different L1 Tracker. The default one is loaded automatically.",
            key="tracker_upload",
        )

    if not validator_files:
        st.markdown("---")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown("#### 1. Upload Submissions")
            st.markdown("Drag in 1 or 2 validator XLSX files from the sidebar.")
        with col2:
            st.markdown("#### 2. Review Verdicts")
            st.markdown("See agreements, divergences, and engine error candidates at a glance.")
        with col3:
            st.markdown("#### 3. Log to L1 Tracker")
            st.markdown("Optionally upload your L1 Feedback Tracker to auto-populate it.")
        st.markdown("---")
        st.info("Upload validator submission(s) in the sidebar to get started.")
        return

    if len(validator_files) > 2:
        st.error("Please upload a maximum of 2 validator files.")
        return

    # ââ Parse uploads ââ
    parsed = []
    for vf in validator_files:
        with st.spinner(f"Parsing {vf.name}..."):
            try:
                data = parse_validator_xlsx(vf.read(), vf.name)
                parsed.append(data)
                st.success(f"Parsed **{vf.name}**: {len(data['boq_comparison'])} BOQ items, "
                          f"{len(data['missing_from_boq'])} missing items, "
                          f"{len(data['discoveries'])} discoveries")
            except Exception as e:
                st.error(f"Error parsing {vf.name}: {str(e)}")
                return

    st.markdown("---")

    # ââ Single validator view ââ
    if len(parsed) == 1:
        v = parsed[0]
        st.markdown("### Validator Submission Summary")

        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("BOQ Comparison Items", len(v['boq_comparison']))
        with col2:
            st.metric("Missing from BOQ Items", len(v['missing_from_boq']))
        with col3:
            st.metric("Validator Discoveries", len(v['discoveries']))

        # Agree/Disagree breakdown
        agrees = sum(1 for i in v['boq_comparison'] if i['agree'] in ('YES', 'Y'))
        disagrees = sum(1 for i in v['boq_comparison'] if i['agree'] in ('NO', 'N'))
        st.markdown(f"**BOQ Comparison:** {agrees} agreed, {disagrees} disagreed out of {len(v['boq_comparison'])} items")

        if v['boq_comparison']:
            st.markdown("#### BOQ Comparison Verdicts")
            st.dataframe(
                [{
                    'Equipment': i['equipment'],
                    'BOQ Qty': i['boq_qty'],
                    'Engine Count': i['engine_count'],
                    'Agree?': i['agree'],
                    'Comments': i['comments'],
                } for i in v['boq_comparison']],
                use_container_width=True, hide_index=True,
            )

        if v['missing_from_boq']:
            st.markdown("#### Missing from BOQ Verdicts")
            st.dataframe(
                [{
                    'Equipment': i['equipment'],
                    'Engine Count': i['engine_count'],
                    'Agree?': i['agree'],
                    'Comments': i['comments'],
                } for i in v['missing_from_boq']],
                use_container_width=True, hide_index=True,
            )

        if v['discoveries']:
            st.markdown("#### Validator Discoveries")
            st.dataframe(
                [{
                    'Equipment': i['equipment'],
                    'Unit': i['unit'],
                    'Count': i['count'],
                    'Location': i['location'],
                    'Why Engine Missed': i['reason_missed'],
                } for i in v['discoveries']],
                use_container_width=True, hide_index=True,
            )

    # ââ Dual validator comparison ââ
    else:
        st.markdown("### Dual Validator Comparison")
        st.markdown(f"**Validator 1:** {parsed[0]['filename']}  |  **Validator 2:** {parsed[1]['filename']}")

        comparison = compare_validators(parsed[0], parsed[1])

        # Summary metrics
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Agreements", len(comparison['agreements']))
        with col2:
            st.metric("Divergences", len(comparison['divergences']),
                      delta=f"{len(comparison['divergences'])} need review" if comparison['divergences'] else None,
                      delta_color="inverse")
        with col3:
            st.metric("Engine Error Candidates", len(comparison['engine_errors']),
                      delta_color="inverse")
        with col4:
            st.metric("Discoveries", len(comparison['discoveries']))

        # Divergences first (most actionable)
        if comparison['divergences']:
            st.markdown("#### Divergences (Validators Disagree)")
            st.markdown("_These items need follow-up. Conservative rule: if unsure, INCLUDE the item._")
            st.dataframe(
                [{
                    'Equipment': d['equipment'],
                    'V1 Verdict': d['v1_agree'],
                    'V2 Verdict': d['v2_agree'],
                    'V1 Comments': d['v1_comments'],
                    'V2 Comments': d['v2_comments'],
                } for d in comparison['divergences']],
                use_container_width=True, hide_index=True,
            )

        if comparison['engine_errors']:
            st.markdown("#### Engine Error Candidates (Both Say NO)")
            st.markdown("_Both validators flagged these as wrong. Feed back to engine._")
            st.dataframe(
                [{
                    'Equipment': e['equipment'],
                    'Issue': e['issue'],
                    'V1 Comments': e['v1_comments'],
                    'V2 Comments': e['v2_comments'],
                } for e in comparison['engine_errors']],
                use_container_width=True, hide_index=True,
            )

        if comparison['agreements']:
            with st.expander(f"Agreements ({len(comparison['agreements'])} items)", expanded=False):
                st.dataframe(
                    [{
                        'Equipment': a['equipment'],
                        'Verdict': a['v1_agree'],
                        'V1 Comments': a['v1_comments'],
                        'V2 Comments': a['v2_comments'],
                    } for a in comparison['agreements']],
                    use_container_width=True, hide_index=True,
                )

        if comparison['discoveries']:
            st.markdown("#### All Validator Discoveries")
            st.dataframe(
                [{
                    'Equipment': d['equipment'],
                    'Unit': d['unit'],
                    'Count': d['count'],
                    'Location': d['location'],
                    'Why Missed': d['reason_missed'],
                    'Source': d['source_validator'],
                } for d in comparison['discoveries']],
                use_container_width=True, hide_index=True,
            )

    # ââ L1 Tracker population ââ
    st.markdown("---")
    st.markdown("### L1 Feedback Tracker")

    # Auto-load default tracker from repo, allow override via upload
    tracker_bytes = None
    if tracker_file:
        tracker_bytes = tracker_file.read()
        st.caption("Using uploaded L1 Tracker (override).")
    else:
        # Try to load default tracker from same directory as the app
        default_tracker_path = os.path.join(os.path.dirname(__file__), "TraceQ L1 Feedback Tracker.xlsx")
        if os.path.exists(default_tracker_path):
            with open(default_tracker_path, "rb") as f:
                tracker_bytes = f.read()
            st.caption("Using default L1 Feedback Tracker from repo.")
        else:
            st.info("No L1 Tracker found. Upload one in the sidebar, or add `TraceQ L1 Feedback Tracker.xlsx` to the repo.")

    if tracker_bytes:
        comparison_data = compare_validators(parsed[0], parsed[1]) if len(parsed) == 2 else None

        if st.button("Populate L1 Tracker", type="primary"):
            with st.spinner("Writing to L1 Feedback Tracker..."):
                try:
                    updated = write_to_l1_tracker(tracker_bytes, job_id, parsed, comparison_data)
                    st.success("L1 Feedback Tracker populated successfully!")
                    st.download_button(
                        label="Download Updated L1 Tracker",
                        data=updated,
                        file_name=f"TraceQ_L1_Feedback_Tracker_{job_id}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                except Exception as e:
                    st.error(f"Error writing to tracker: {str(e)}")


# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
# PAGE ROUTER
# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ

if page == "Upload Validator Response":
    render_validator_page()

elif page == "Generate Client Report":
    render_client_report_page()

else:
    # âââ ENGINE ANALYSIS PAGE (original) âââââââââââââââââââââââââââââââââââââ
    with st.sidebar:
        st.markdown("### Upload Files")
        st.markdown("Upload your HVAC drawing(s) to analyse.")

    drawing_files = st.file_uploader(
        "ð Drawing File(s) (DXF or DWG)",
        type=["dxf", "dwg"],
        accept_multiple_files=True,
        help="Upload one or more HVAC layout drawings in DXF or DWG format."
    )

    boq_file = st.file_uploader(
        "ð BOQ Spreadsheet (optional)",
        type=["xlsx", "xls", "csv"],
        help="Upload the Bill of Quantities for comparison. If not provided, TraceQ will still count all equipment found in the drawing."
    )

    st.markdown("---")
    st.markdown("### About")
    st.markdown(
        "TraceQ analyses HVAC drawings and compares equipment counts "
        "against the BOQ to identify discrepancies and missing items."
    )
    st.markdown(
        "**Three-tier detection:**\n"
        "1. Layer-based classification\n"
        "2. Block name matching\n"
        "3. Text label analysis"
    )
    st.markdown("---")
    st.markdown("### DWG Support")
    st.markdown(
        "â **DWG files are supported.** Upload a DWG directly and "
        "TraceQ will convert it to DXF automatically on the server."
    )
    st.markdown(
        "_If auto-conversion fails, you can also convert manually:_\n"
        "- **AutoCAD/BricsCAD**: File â Save As â DXF\n"
        "- **Online**: [CloudConvert](https://cloudconvert.com/dwg-to-dxf)"
    )
    st.markdown("---")
    st.markdown("*Built by [TechTelligence](mailto:nicholas@ttelligence.com)*")
    st.markdown("*v1.3 â March 2026*")


    # âââ Main Content âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ

    if not drawing_files:
        # Landing state â no file uploaded yet
        st.markdown("---")

        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown("#### ð Upload Drawing(s)")
            st.markdown("Upload your HVAC layout drawing(s) (DXF or DWG) using the sidebar.")
        with col2:
            st.markdown("#### ð Automatic Analysis")
            st.markdown("TraceQ scans every layer, block, and text label to count equipment.")
        with col3:
            st.markdown("#### ð Get Your Report")
            st.markdown("See discrepancies, missing items, and cost exposure at a glance.")

        st.markdown("---")
        st.info("ð Upload one or more DXF/DWG files in the sidebar to get started.")

    else:
        # âââ Prepare all uploaded drawing files âââââââââââââââââââââââââââââââââââ
        tmp_paths = []  # List of (filename, tmp_path) tuples
        for drawing_file in drawing_files:
            file_ext = os.path.splitext(drawing_file.name)[1].lower() or '.dxf'
            with tempfile.NamedTemporaryFile(suffix=file_ext, delete=False) as tmp:
                tmp.write(drawing_file.read())
                tmp_path = tmp.name

            # DWG â DXF Auto-Conversion
            if file_ext == '.dwg':
                with st.spinner(f"Converting {drawing_file.name} DWG to DXF..."):
                    try:
                        dxf_path = FileConverter.convert_dwg_to_dxf(tmp_path)
                        tmp_path = dxf_path
                        st.success(f"â Converted **{drawing_file.name}** to DXF successfully.")
                    except RuntimeError as e:
                        st.error(
                            f"â ï¸ Could not convert {drawing_file.name} automatically.\n\n"
                            f"**What to do:** Open the DWG in AutoCAD or BricsCAD â File â Save As â DXF, "
                            f"then upload the DXF version.\n\n"
                            f"_Technical detail: {str(e)}_"
                        )
                        continue  # Skip this file, process the rest

            tmp_paths.append((drawing_file.name, tmp_path))

        if not tmp_paths:
            st.error("No valid drawing files to process.")
            st.stop()

        # âââ Display file count âââââââââââââââââââââââââââââââââââââââââââââââââââ
        drawing_names = [name for name, _ in tmp_paths]
        drawing_name_combined = " + ".join(drawing_names)
        if len(tmp_paths) > 1:
            st.info(f"ð **{len(tmp_paths)} drawing files** uploaded for combined analysis.")

        # âââ Run Quick Scan (shared between tabs) ââââââââââââââââââââââââââââââââââ
        # Quick Scan tab shows first file. For multi-file feedback sheet, merge all scans.
        scan = None
        scan_for_feedback = None
        with st.spinner("Running quick scan..."):
            try:
                engine = TraceQEngine()
                scan = engine.quick_scan(tmp_paths[0][1])
                # For multi-file: merge unrecognised blocks/layers from ALL files
                if len(tmp_paths) > 1:
                    import copy
                    scan_for_feedback = copy.deepcopy(scan)
                    all_unrec_blocks = {}
                    all_unrec_layers = set()
                    for _, fpath in tmp_paths:
                        fs = engine.quick_scan(fpath)
                        for ub in (getattr(fs, 'unrecognised_blocks', []) or []):
                            bname = ub.get('block', ub) if isinstance(ub, dict) else str(ub)
                            cnt = ub.get('count', 0) if isinstance(ub, dict) else 0
                            if bname in all_unrec_blocks:
                                all_unrec_blocks[bname]['count'] += cnt
                            else:
                                all_unrec_blocks[bname] = {'block': bname, 'count': cnt}
                        for ul in (getattr(fs, 'unrecognised_layers', []) or []):
                            layer_name = ul if isinstance(ul, str) else str(ul)
                            all_unrec_layers.add(layer_name)
                    scan_for_feedback.unrecognised_blocks = list(all_unrec_blocks.values())
                    scan_for_feedback.unrecognised_layers = list(all_unrec_layers)
                else:
                    scan_for_feedback = scan
            except Exception as e:
                st.error(f"Quick scan failed: {str(e)}")

        # âââ Step 0: Quick Scan + Full Analysis Tabs ââââââââââââââââââââââââââââââ
        tab_scan, tab_analysis = st.tabs(["Step 0: Quick Scan", "Full Analysis"])

        # âââ TAB 1: QUICK SCAN âââââââââââââââââââââââââââââââââââââââââââââââââââ
        with tab_scan:
            st.markdown("### Step 0 â Compatibility Scan")
            st.caption("Quick check: how much of this drawing does TraceQ recognise?")

            if scan and scan._dwg_unsupported:
                st.error(scan.verdict_msg)
            elif scan:
                # ââ Overall Score ââ
                if scan.verdict == 'HIGH':
                    score_color = "ð¢"
                    st.success(f"{score_color} **Overall Compatibility: {scan.overall_score}% â HIGH**")
                    st.info(scan.verdict_msg)
                elif scan.verdict == 'MEDIUM':
                    score_color = "ð¡"
                    st.warning(f"{score_color} **Overall Compatibility: {scan.overall_score}% â MEDIUM**")
                    st.info(scan.verdict_msg)
                else:
                    score_color = "ð´"
                    st.error(f"{score_color} **Overall Compatibility: {scan.overall_score}% â LOW**")
                    st.info(scan.verdict_msg)

                # ââ Score Breakdown ââ
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Layers", f"{scan.layer_score}%",
                              delta=f"{len(scan.recognised_layers)}/{scan.hvac_candidate_layers}")
                with col2:
                    st.metric("Blocks", f"{scan.block_score}%",
                              delta=f"{len(scan.recognised_blocks)}/{scan.total_blocks}")
                with col3:
                    st.metric("Text Patterns", f"{scan.mtext_score}%",
                              delta=f"{scan.mtext_pattern_hits}/{scan.total_mtext_patterns}")
                with col4:
                    st.metric("Total Entities", f"{scan.total_entities:,}")

                st.markdown("---")

                # ââ Recognised Layers ââ
                if scan.recognised_layers:
                    with st.expander(f"â Recognised Layers ({len(scan.recognised_layers)})", expanded=True):
                        layer_data = []
                        for rl in scan.recognised_layers:
                            layer_data.append({
                                "Layer Name": rl['layer'],
                                "Equipment Type": rl['equipment_type'].replace('_', ' ').title(),
                                "Confidence": f"{rl['confidence']:.0%}",
                                "Match": rl['method'],
                            })
                        st.dataframe(layer_data, use_container_width=True, hide_index=True)

                # ââ Unrecognised Layers ââ
                if scan.unrecognised_layers:
                    with st.expander(f"â Unrecognised Layers ({len(scan.unrecognised_layers)})", expanded=False):
                        st.caption("These layers may contain equipment that TraceQ doesn't recognise yet. Nestor can help identify them.")
                        for ul in scan.unrecognised_layers:
                            st.text(f"  {ul}")

                # ââ Recognised Blocks ââ
                if scan.recognised_blocks:
                    with st.expander(f"â Recognised Blocks ({len(scan.recognised_blocks)})", expanded=True):
                        block_data = []
                        for rb in scan.recognised_blocks:
                            block_data.append({
                                "Block Name": rb['block'],
                                "Equipment Type": rb['equipment_type'].replace('_', ' ').title(),
                                "Count": rb['count'],
                                "Match": rb['match'],
                            })
                        st.dataframe(block_data, use_container_width=True, hide_index=True)

                # ââ Unrecognised Blocks ââ
                if scan.unrecognised_blocks:
                    with st.expander(f"â Unrecognised Blocks ({len(scan.unrecognised_blocks)})", expanded=False):
                        st.caption("These blocks may be equipment. Nestor can identify them to expand the dictionary.")
                        block_unk = []
                        for ub in scan.unrecognised_blocks:
                            block_unk.append({
                                "Block Name": ub['block'],
                                "Occurrences": ub['count'],
                            })
                        st.dataframe(block_unk, use_container_width=True, hide_index=True)

                st.markdown("---")
                st.caption("Tip: After running the full analysis, send unrecognised items to Nestor for identification. His corrections will permanently improve TraceQ's accuracy.")

        # âââ TAB 2: FULL ANALYSIS ââââââââââââââââââââââââââââââââââââââââââââââââ
        with tab_analysis:
            # âââ Analyse all drawing files with multi-view dedup âââââââââââââââââââââ
            all_results = []
            combined_merged = {}
            combined_parse_info = {'layers': 0, 'block_types': 0}
            combined_dedup_report = None

            with st.spinner(f"Analysing {'drawings' if len(tmp_paths) > 1 else 'drawing'}... this may take a moment."):
                engine = TraceQEngine()

                # Use analyze_multi for multi-view dedup + non-layout file filtering
                file_paths = [fpath for _, fpath in tmp_paths]
                multi_result = engine.analyze_multi(file_paths)

                all_results = multi_result['results']
                combined_merged = multi_result['combined']
                skipped_files = multi_result.get('skipped', [])
                floor_groups = multi_result.get('floor_groups', [])

                # Merge parse info from all results
                for fname, result in all_results:
                    p = result.parse_info
                    combined_parse_info['layers'] += p.get('layers', 0)
                    combined_parse_info['block_types'] += p.get('block_types', 0)

                # Use the last result's dedup report (proximity dedup within single files)
                if all_results:
                    result = all_results[-1][1]
                    combined_dedup_report = getattr(result, 'dedup_report', None)

            # Clean up temp files
            for _, fpath in tmp_paths:
                try:
                    os.unlink(fpath)
                except OSError:
                    pass

            if not all_results:
                st.error("No drawing files could be analysed.")
                st.stop()

            # âââ Results Header âââââââââââââââââââââââââââââââââââââââââââââââââââââââ
            if len(all_results) == 1:
                st.success(f"â Analysis complete â **{all_results[0][0]}**")
            else:
                msg = f"â Analysis complete â **{len(all_results)} files** combined"
                if skipped_files:
                    msg += f" ({len(skipped_files)} non-layout files filtered)"
                has_multi_view = any(len(g) > 1 for g in floor_groups)
                if has_multi_view:
                    msg += " with multi-view deduplication"
                st.success(msg)
            st.markdown("---")

            # âââ Key Metrics ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
            merged = combined_merged
            total_items = sum(v.get('count', 0) for v in merged.values())
            total_categories = len(merged)
            parser_info = combined_parse_info
            _dname = drawing_names[0].split('.')[0] if len(drawing_names) == 1 else f"{len(drawing_names)}_files"

            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total Equipment", f"{total_items:,}")
            with col2:
                st.metric("Categories Found", total_categories)
            with col3:
                st.metric("Layers Scanned", parser_info.get('layers', 0))
            with col4:
                st.metric("Block Types", parser_info.get('block_types', 0))

            st.markdown("---")

            # âââ Equipment Inventory ââââââââââââââââââââââââââââââââââââââââââââââââââ
            st.markdown("### ð Equipment Inventory")

            table_data = []
            review_items = []
            for equip_type, data in sorted(merged.items()):
                count = data.get('count', 0)
                source = data.get('source', 'unknown')
                confidence = data.get('confidence', 0)
                alt = data.get('alternate_counts', {})
                flagged = data.get('needs_review', False)

                if 'tier1' in source:
                    source_label = "ð¢ Layer"
                elif 'tier2' in source:
                    source_label = "ðµ Block"
                elif 'tier3' in source:
                    source_label = "ð¡ Text"
                else:
                    source_label = f"âª {source}"

                name = _format_equipment_name(equip_type)

                # Show all three tier counts explicitly
                t1 = alt.get('tier1', 0)
                t2 = alt.get('tier2', 0)
                t3 = alt.get('tier3', 0)

                row = {
                    "Equipment": name,
                    "Count": count,
                    "Source": source_label,
                    "Confidence": f"{int(confidence * 100)}%",
                    "Layer": t1 if t1 > 0 else "â",
                    "Block": t2 if t2 > 0 else "â",
                    "Text": t3 if t3 > 0 else "â",
                }

                if flagged:
                    row["Flag"] = "â ï¸ Review"
                    review_items.append({
                        'name': name,
                        'note': data.get('notes', 'Tier counts disagree significantly.'),
                        'tier1': t1, 'tier2': t2, 'tier3': t3,
                    })
                else:
                    row["Flag"] = "â"

                table_data.append(row)

            if table_data:
                st.dataframe(
                    table_data,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "Equipment": st.column_config.TextColumn("Equipment", width="medium"),
                        "Count": st.column_config.NumberColumn("Count", width="small"),
                        "Source": st.column_config.TextColumn("Source", width="small"),
                        "Confidence": st.column_config.TextColumn("Conf.", width="small"),
                        "Layer": st.column_config.TextColumn("Layer (T1)", width="small"),
                        "Block": st.column_config.TextColumn("Block (T2)", width="small"),
                        "Text": st.column_config.TextColumn("Text (T3)", width="small"),
                        "Flag": st.column_config.TextColumn("Status", width="small"),
                    }
                )

            # Show review warnings if any
            if review_items:
                st.markdown("#### â ï¸ Items Flagged for QS Review")
                for item in review_items:
                    st.warning(
                        f"**{item['name']}** â Tier counts disagree: "
                        f"Layer={item['tier1']}, Block={item['tier2']}, Text={item['tier3']}. "
                        f"Recommend manual verification."
                    )

            # Show dedup report if any proximity deductions were made
            dedup_report = result.detection_results.get('dedup_report', {})
            if dedup_report:
                adjustments = dedup_report.get('adjustments', [])
                if adjustments:
                    with st.expander(f"ð Proximity Deduplication ({len(adjustments)} adjustments)", expanded=False):
                        st.caption(
                            "Text labels found near block INSERTs of the same equipment type â "
                            "Tier 3 count reduced to avoid double-counting."
                        )
                        for adj in adjustments:
                            st.info(
                                f"**{_format_equipment_name(adj.get('equipment_type', ''))}** â "
                                f"Tier 3 reduced from {adj.get('tier3_original', 0)} to {adj.get('tier3_adjusted', 0)} "
                                f"({adj.get('shadowed_by_blocks', 0)} text labels near blocks, "
                                f"radius: {dedup_report.get('radius_used', 0):.0f} units)"
                            )

            st.markdown("---")

            # âââ BOQ Comparison (if BOQ uploaded) âââââââââââââââââââââââââââââââââââââ
            if boq_file is not None:
                st.markdown("### ð BOQ Discrepancy Report")

                try:
                    boq_bytes = boq_file.read()
                    boq_items = parse_boq(boq_bytes, boq_file.name)

                    if boq_items:
                        # ââ BOQ Coverage Check (Tier 1 pre-flight) ââ
                        classified, unclassified = boq_coverage_check(boq_items)
                        coverage_pct = len(classified) / len(boq_items) * 100 if boq_items else 0

                        with st.expander(
                            f"ð BOQ Coverage Check â {len(classified)}/{len(boq_items)} items classified ({coverage_pct:.0f}%)",
                            expanded=len(unclassified) > 0
                        ):
                            if unclassified:
                                st.warning(
                                    f"**{len(unclassified)} BOQ line item(s) could not be classified** and will be "
                                    f"excluded from the comparison. These items have no matching equipment type in "
                                    f"the synonym library."
                                )
                                import pandas as pd
                                uncl_df = pd.DataFrame([
                                    {
                                        'Description': u['description'],
                                        'Qty': u['qty'],
                                        'Unit': u['unit'],
                                    }
                                    for u in unclassified
                                ])
                                st.dataframe(uncl_df, use_container_width=True, hide_index=True)
                                st.caption(
                                    "These items may need manual review or a Tier 2 RFI. "
                                    "Future engine updates will expand the synonym library to cover more description patterns."
                                )
                            else:
                                st.success("All BOQ line items classified successfully. Full coverage achieved.")

                            if classified:
                                # Show classification summary
                                from collections import Counter
                                type_counts = Counter(c['equipment_type'] for c in classified)
                                st.caption(
                                    f"Classified into {len(type_counts)} equipment types: "
                                    + ", ".join(f"{_format_equipment_name(t)} ({n})" for t, n in type_counts.most_common())
                                )

                        comparisons, missing_from_boq = compare_boq_vs_drawing(boq_items, merged)

                        # ââ Summary Metrics ââ
                        matches = sum(1 for c in comparisons if c['Risk'] == 'MATCH')
                        discrepancies = sum(1 for c in comparisons if c['Risk'] == 'DISCREPANCY')
                        missing_count = len(missing_from_boq)
                        comparison_exp = sum(c.get('_exposure_num') or 0 for c in comparisons)
                        missing_exp = sum(m.get('_est_exposure') or 0 for m in missing_from_boq)
                        total_exposure = comparison_exp + missing_exp

                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            st.metric("Matched", matches)
                        with col2:
                            st.metric("Discrepancies", discrepancies)
                        with col3:
                            st.metric("Missing from BOQ", missing_count)
                        with col4:
                            st.metric("Total Exposure", f"AED {total_exposure:,.0f}")

                        # ââ EXCEL DOWNLOAD â top of report ââ
                        excel_bytes = generate_excel_report(
                            comparisons, missing_from_boq, boq_items,
                            drawing_name_combined, boq_file.name,
                            merged=merged,
                            dedup_report=result.detection_results.get('dedup_report'),
                        )
                        report_filename = f"TraceQ_BOQ_Report_{_dname}_{datetime.now().strftime('%Y%m%d')}.xlsx"

                        col_dl1, col_dl2 = st.columns(2)
                        with col_dl1:
                            st.download_button(
                                label="ð¥ Download BOQ Report (Client)",
                                data=excel_bytes,
                                file_name=report_filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                type="primary",
                            )
                        with col_dl2:
                            validator_bytes = generate_validator_template(
                                comparisons, missing_from_boq, merged,
                                drawing_name_combined,
                                scan=scan_for_feedback,
                            )
                            validator_filename = f"TraceQ_Validator_{_dname}_{datetime.now().strftime('%Y%m%d')}.xlsx"
                            st.download_button(
                                label="ð Download Validator Template",
                                data=validator_bytes,
                                file_name=validator_filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            )

                        st.markdown("---")

                        # ââ Main Comparison Table ââ
                        st.markdown("#### Comparison Details")

                        display_comparisons = []
                        for c in comparisons:
                            display_comparisons.append({
                                'Trace ID': c['Trace ID'],
                                'Equipment': c['Equipment'],
                                'BOQ Qty': c['BOQ Qty'],
                                'Drawing Qty': c['Drawing Qty'],
                                'Diff': c['Difference'],
                                'Unit': c['Unit'],
                                'Status': c['Risk'],
                                'Exposure (AED)': c['Exposure (AED)'],
                                'Notes': c['Notes'],
                            })

                        st.dataframe(
                            display_comparisons,
                            use_container_width=True,
                            hide_index=True,
                            column_config={
                                "Trace ID": st.column_config.TextColumn("Trace ID", width="small"),
                                "Equipment": st.column_config.TextColumn("Equipment", width="medium"),
                                "BOQ Qty": st.column_config.TextColumn("BOQ", width="small"),
                                "Drawing Qty": st.column_config.TextColumn("Drawing", width="small"),
                                "Diff": st.column_config.TextColumn("Diff", width="small"),
                                "Unit": st.column_config.TextColumn("Unit", width="small"),
                                "Status": st.column_config.TextColumn("Status", width="small"),
                                "Exposure (AED)": st.column_config.TextColumn("Exposure", width="small"),
                                "Notes": st.column_config.TextColumn("Notes", width="large"),
                            }
                        )

                        # ââ Missing from BOQ ââ
                        if missing_from_boq:
                            st.markdown(f"#### Items in Drawing Not in BOQ ({missing_count} items)")
                            st.caption("These items were detected in the drawing but have no corresponding BOQ line item.")

                            missing_display = []
                            for m in missing_from_boq:
                                missing_display.append({
                                    'Trace ID': m['Trace ID'],
                                    'Equipment': m['Equipment'],
                                    'Drawing Qty': m['Drawing Qty'],
                                    'Detection': m['Detection'],
                                    'Confidence': m['Confidence'],
                                    'Notes': m['Notes'],
                                })

                            st.dataframe(
                                missing_display,
                                use_container_width=True,
                                hide_index=True,
                                column_config={
                                    "Trace ID": st.column_config.TextColumn("Trace ID", width="small"),
                                    "Equipment": st.column_config.TextColumn("Equipment", width="medium"),
                                    "Drawing Qty": st.column_config.NumberColumn("Qty", width="small"),
                                    "Detection": st.column_config.TextColumn("Detection", width="medium"),
                                    "Confidence": st.column_config.TextColumn("Confidence", width="small"),
                                    "Notes": st.column_config.TextColumn("Notes", width="large"),
                                }
                            )

                        # ââ Parsed BOQ Line Items (detail expander) ââ
                        with st.expander("ð Parsed BOQ Line Items", expanded=False):
                            boq_display = []
                            for item in boq_items:
                                boq_display.append({
                                    "Ref": item.get('boq_ref', 'â'),
                                    "Description": item['description'][:70],
                                    "Type": (item['equipment_type'] or 'â').replace('_', ' ').title(),
                                    "Unit": item.get('unit', 'â'),
                                    "Qty": int(item['qty']) if item['qty'] == int(item['qty']) else item['qty'],
                                    "Rate": f"{item['rate']:,.0f}" if item.get('rate') else 'â',
                                    "Total": f"{item['total']:,.0f}" if item.get('total') else 'â',
                                })
                            st.dataframe(boq_display, use_container_width=True, hide_index=True)
                    else:
                        st.warning("Could not parse any equipment items from the BOQ file. Check the format.")

                except Exception as e:
                    st.error(f"Error reading BOQ file: {str(e)}")

                st.markdown("---")

            # âââ Validation Results âââââââââââââââââââââââââââââââââââââââââââââââââââ
            st.markdown("### â ï¸ Validation Checks")

            validation = result.validation_results
            warnings = validation.get('warnings', [])

            if not warnings:
                st.success("All validation checks passed â no warnings.")
            else:
                for w in warnings:
                    if isinstance(w, dict):
                        severity = w.get('severity', 'info')
                        msg = w.get('message', str(w))
                        if severity == 'warning':
                            st.warning(msg)
                        elif severity == 'critical':
                            st.error(msg)
                        else:
                            st.info(msg)
                    else:
                        w_str = str(w)
                        if w_str.startswith('[WARNING]'):
                            st.warning(w_str)
                        elif w_str.startswith('[CRITICAL]'):
                            st.error(w_str)
                        else:
                            st.info(w_str)

            st.markdown("---")

            # âââ Layer Classification âââââââââââââââââââââââââââââââââââââââââââââââââ
            with st.expander("ðï¸ Layer Classification Details", expanded=False):
                layer_results = result.layer_classification
                classified = []
                unclassified = []

                for layer_name, info in sorted(layer_results.items()):
                    equip = info.get('equipment_type')
                    conf = info.get('confidence', 0)
                    method = info.get('method', 'unknown')

                    if equip:
                        classified.append({
                            "Layer": layer_name,
                            "Equipment Type": equip.replace('_', ' ').title(),
                            "Confidence": f"{int(conf * 100)}%",
                            "Match": method,
                        })
                    else:
                        unclassified.append(layer_name)

                if classified:
                    st.markdown("**Classified Layers:**")
                    st.dataframe(classified, use_container_width=True, hide_index=True)

                if unclassified:
                    st.markdown(f"**Unclassified Layers ({len(unclassified)}):**")
                    st.text(", ".join(unclassified))

            # âââ Detection Tier Breakdown âââââââââââââââââââââââââââââââââââââââââââââ
            with st.expander("ð Detection Tier Breakdown", expanded=False):
                for tier_name in ['tier1', 'tier2', 'tier3']:
                    tier_data = result.detection_results.get(tier_name, {})
                    if tier_data:
                        labels = {'tier1': 'ð¢ Tier 1 â Layer Detection',
                                  'tier2': 'ðµ Tier 2 â Block Detection',
                                  'tier3': 'ð¡ Tier 3 â Text Detection'}
                        st.markdown(f"**{labels[tier_name]}**")
                        tier_items = []
                        for equip, data in sorted(tier_data.items()):
                            tier_items.append({
                                "Equipment": equip.replace('_', ' ').title(),
                                "Count": data.get('count', 0),
                            })
                        st.dataframe(tier_items, use_container_width=True, hide_index=True)

            # âââ Raw JSON Output ââââââââââââââââââââââââââââââââââââââââââââââââââââââ
            with st.expander("ð§ Raw JSON Output", expanded=False):
                st.json(result.to_dict())

            # âââ Download Button (JSON fallback â always available) âââââââââââââââââââ
            st.markdown("---")
            json_output = json.dumps(result.to_dict(), indent=2)
            st.download_button(
                label="ð¥ Download Full Analysis (JSON)",
                data=json_output,
                file_name=f"TraceQ_Analysis_{_dname}_{datetime.now().strftime('%Y%m%d')}.json",
                mime="application/json",
            )
