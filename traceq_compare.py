#!/usr/bin/env python3
"""
TraceQ Compare Module — Standalone BOQ vs Drawing Comparison
=============================================================
Extracted from streamlit_app.py to run WITHOUT streamlit dependency.
This is the SINGLE SOURCE OF TRUTH for comparison logic.
Both the Streamlit app and the PDF generator must use this module.

Usage:
    from traceq_compare import parse_boq_file, compare_boq_vs_drawing

    boq_items = parse_boq_file("path/to/boq.xlsx")
    engine = TraceQEngine()
    result = engine.analyze("path/to/drawing.dxf")
    comparisons, missing = compare_boq_vs_drawing(boq_items, result.merged)

Built by TechTelligence — nicholas@ttelligence.com
"""

import os
import re
import openpyxl

# ═══════════════════════════════════════════════════════════════════
# CONFIGURATION — shared constants (single source of truth)
# ═══════════════════════════════════════════════════════════════════

# Trace ID prefix map: equipment_type → category prefix for TQ-[CAT]-[NNN] format
TRACE_PREFIX_MAP = {
    'supply_duct': 'DUCT', 'return_duct': 'DUCT', 'exhaust_duct': 'DUCT',
    'volume_control_damper': 'VCD', 'fcu': 'FCU',
    'supply_diffuser': 'DIFF', 'return_diffuser': 'DIFF', 'extract_diffuser': 'DIFF',
    'flow_bar': 'FLOW', 'thermostat': 'THERM', 'vrf': 'VRF',
    'flexible_duct': 'FLEX', 'plenum_box': 'PLEN',
    'drain_pipe': 'PIPE', 'refrigerant_pipe': 'PIPE',
    'sound_attenuator': 'ACOU', 'fire_damper': 'DAMP', 'motorized_damper': 'DAMP',
    'non_return_damper': 'DAMP', 'indoor_unit': 'FCU', 'outdoor_unit': 'VRF',
    'exhaust_fan': 'FAN', 'grille': 'DIFF', 'insulation': 'MISC',
    'access_door': 'MISC', 'damper_general': 'DAMP', 'hvac_equipment': 'EQUIP',
}

# Estimated UAE HVAC unit rates (AED) for missing-from-BOQ exposure calculation
UAE_UNIT_RATES = {
    'fcu': 2500, 'supply_diffuser': 180, 'return_diffuser': 180,
    'extract_diffuser': 200, 'volume_control_damper': 350, 'thermostat': 150,
    'vrf': 45000, 'flow_bar': 120, 'plenum_box': 280, 'indoor_unit': 2500,
    'outdoor_unit': 55000, 'fire_damper': 450, 'motorized_damper': 800,
    'sound_attenuator': 600, 'exhaust_fan': 3500, 'flexible_duct': 85,
}

# Expected detection method per equipment type — used in Trace ID reference
EXPECTED_DETECTION_METHOD = {
    'supply_duct': 'Layer Detection', 'return_duct': 'Layer Detection', 'exhaust_duct': 'Layer Detection',
    'volume_control_damper': 'Layer Detection', 'fcu': 'Block Detection',
    'supply_diffuser': 'Layer Detection', 'return_diffuser': 'Block Detection',
    'extract_diffuser': 'Text/Label Detection', 'thermostat': 'Layer Detection',
    'flow_bar': 'Text/Label Detection', 'vrf': 'Text/Label Detection',
    'flexible_duct': 'Layer Detection', 'plenum_box': 'Text/Label Detection',
    'indoor_unit': 'Block Detection', 'outdoor_unit': 'Block Detection',
    'fire_damper': 'Block Detection', 'motorized_damper': 'Block Detection',
    'non_return_damper': 'Block Detection', 'sound_attenuator': 'Block Detection',
    'exhaust_fan': 'Block Detection', 'grille': 'Block Detection',
    'insulation': 'Layer Detection', 'access_door': 'Block Detection',
    'drain_pipe': 'Layer Detection', 'refrigerant_pipe': 'Layer Detection',
}

# BOQ keyword map for classifying BOQ descriptions → engine equipment types
BOQ_KEYWORD_MAP = [
    ('FCU-1', 'fcu', 'FCU-1 Ducted'),
    ('FCU-2', 'fcu', 'FCU-2 Ducted'),
    ('FCU-3', 'fcu', 'FCU-3 Ducted'),
    ('FCU-4', 'fcu', 'FCU-4'),
    ('FCU', 'fcu', 'FCU (General)'),
    ('FAN COIL', 'fcu', 'Fan Coil Unit'),
    ('THERMOSTAT', 'thermostat', 'Thermostat'),
    ('SUPPLY AIR DIFFUSER', 'supply_diffuser', 'Supply Air Diffuser'),
    ('RETURN AIR DIFFUSER', 'return_diffuser', 'Return Air Diffuser'),
    ('SUPPLY AIR FLOW BAR', 'flow_bar', 'Supply Air Flow Bar'),
    ('RETURN AIR FLOW BAR', 'flow_bar', 'Return Air Flow Bar'),
    ('FLOW BAR', 'flow_bar', 'Flow Bar'),
    ('PLENUM BOX', 'plenum_box', 'Plenum Box'),
    ('SUPPLY AIR VOLUME DAMPER', 'volume_control_damper', 'Supply Air Volume Damper'),
    ('VOLUME DAMPER', 'volume_control_damper', 'Volume Control Damper'),
    ('VCD', 'volume_control_damper', 'VCD'),
    ('FIRE DAMPER', 'fire_damper', 'Fire Damper'),
    ('MOTORIZED DAMPER', 'motorized_damper', 'Motorized Damper'),
    ('NON RETURN DAMPER', 'non_return_damper', 'Non-Return Damper'),
    ('SOUND ATTENUATOR', 'sound_attenuator', 'Sound Attenuator'),
    ('SUPPLY AIR DUCT', 'supply_duct', 'Supply Air Duct'),
    ('RETURN AIR DUCT', 'return_duct', 'Return Air Duct'),
    ('FLEXIBLE DUCT', 'flexible_duct', 'Flexible Duct'),
    ('VRV', 'vrf', 'VRV/VRF Unit'),
    ('VRF', 'vrf', 'VRF Unit'),
    ('OUTDOOR UNIT', 'outdoor_unit', 'Outdoor Unit'),
    ('INDOOR UNIT', 'indoor_unit', 'Indoor Unit'),
    ('WALL MOUNTED', 'indoor_unit', 'Wall Mounted Unit'),
    ('GRILLE', 'grille', 'Grille'),
    ('EXHAUST FAN', 'exhaust_fan', 'Exhaust Fan'),
    ('VENTILATION FAN', 'exhaust_fan', 'Ventilation Fan'),
    ('ACCESS DOOR', 'access_door', 'Access Door'),
    ('DRAIN', 'drain_pipe', 'Drain Pipe'),
    ('INSULATION', 'insulation', 'Insulation'),
]


# ═══════════════════════════════════════════════════════════════════
# BOQ PARSER
# ═══════════════════════════════════════════════════════════════════

def _detect_boq_columns(ws):
    """Detect which columns hold description, unit, qty, rate, total."""
    cols = {'desc': 3, 'unit': 4, 'qty': 5, 'rate': 6, 'total': 7, 'item_no': 2}

    for row_num in range(1, min(11, ws.max_row + 1)):
        for col_num in range(1, min(11, ws.max_column + 1)):
            val = ws.cell(row=row_num, column=col_num).value
            if val is None:
                continue
            upper = str(val).strip().upper()
            if upper in ('UNIT', 'UOM', 'U/M'):
                cols['unit'] = col_num
            elif upper in ('QTY', 'QUANTITY', 'QTY.'):
                cols['qty'] = col_num
            elif 'RATE' in upper and 'TOTAL' not in upper:
                cols['rate'] = col_num
            elif 'TOTAL' in upper:
                cols['total'] = col_num

    text_col_counts = {}
    for row_num in range(5, min(20, ws.max_row + 1)):
        for col_num in range(1, min(8, ws.max_column + 1)):
            val = ws.cell(row=row_num, column=col_num).value
            if isinstance(val, str) and len(val.strip()) > 5:
                text_col_counts[col_num] = text_col_counts.get(col_num, 0) + 1

    if text_col_counts:
        best_col = max(text_col_counts, key=text_col_counts.get)
        cols['desc'] = best_col
        cols['item_no'] = best_col - 1 if best_col > 1 else 1

    return cols


def _classify_description(desc_text):
    """Classify a BOQ description into an equipment type."""
    upper = desc_text.upper().strip()
    for keyword, etype, label in BOQ_KEYWORD_MAP:
        if keyword in upper:
            return etype, label
    return None, desc_text


def parse_boq_file(filepath):
    """
    Parse a BOQ Excel file from disk and extract equipment line items.
    Returns list of dicts with keys: description, equipment_type, equipment_label,
    qty, unit, rate, total, boq_ref, is_countable
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    cols = _detect_boq_columns(ws)
    items = []
    item_counter = 0

    for row_num in range(1, ws.max_row + 1):
        item_no_val = ws.cell(row=row_num, column=cols['item_no']).value
        desc_val = ws.cell(row=row_num, column=cols['desc']).value
        unit_val = ws.cell(row=row_num, column=cols['unit']).value
        qty_val = ws.cell(row=row_num, column=cols['qty']).value
        rate_val = ws.cell(row=row_num, column=cols['rate']).value
        total_val = ws.cell(row=row_num, column=cols['total']).value

        if desc_val is None:
            continue
        desc_str = str(desc_val).strip()
        if len(desc_str) < 3:
            continue
        if qty_val is None:
            continue
        try:
            qty = float(qty_val)
        except (ValueError, TypeError):
            continue
        if qty <= 0:
            continue

        equip_type, equip_label = _classify_description(desc_str)
        unit_str = str(unit_val).strip() if unit_val else None
        try:
            rate = float(rate_val) if rate_val else None
        except (ValueError, TypeError):
            rate = None
        try:
            total = float(total_val) if total_val else None
        except (ValueError, TypeError):
            total = None

        item_counter += 1
        boq_ref = str(item_no_val).strip() if item_no_val else str(item_counter)

        items.append({
            'description': desc_str,
            'equipment_type': equip_type,
            'equipment_label': equip_label,
            'qty': qty,
            'unit': unit_str,
            'rate': rate,
            'total': total,
            'boq_ref': boq_ref,
            'is_countable': (unit_str or '').lower().strip('.') in {'nos', 'no', 'pcs', 'ea', 'each', 'set', 'sets'},
        })

    return items


# ═══════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════

def _format_equipment_name(etype):
    """Format equipment type to proper display name, preserving acronyms."""
    acronyms = {'fcu': 'FCU', 'vrf': 'VRF', 'vcd': 'VCD', 'sad': 'SAD', 'rad': 'RAD'}
    name = etype.replace('_', ' ').title()
    for key, acr in acronyms.items():
        name = name.replace(key.title(), acr)
    return name


def _format_source_label(source):
    """Convert raw engine source to clean label."""
    if not source or source == '—':
        return '—'
    if 'tier1' in source:
        return 'Layer Detection'
    elif 'tier2' in source:
        return 'Block Detection'
    elif 'tier3' in source or 'SAD' in source:
        return 'Text/Label Detection'
    return source


def _format_boq_breakdown(items):
    if len(items) == 1:
        return items[0]['description'][:60]
    parts = []
    for item in items:
        desc_short = item['description'][:40]
        parts.append(f"{desc_short}: {int(item['qty']) if item['qty'] == int(item['qty']) else item['qty']}")
    return " | ".join(parts)


def _build_discrepancy_note(etype, boq_data, drawing_data, diff):
    drawing_qty = drawing_data.get('count', 0)
    boq_qty = boq_data['total_qty']
    source = drawing_data.get('source', 'detection')

    if 'tier1' in source:
        source_label = "layer detection"
    elif 'tier2' in source:
        source_label = "block detection"
    elif 'tier3' in source or 'SAD' in source:
        source_label = "text/label detection"
    else:
        source_label = source

    if diff > 0:
        note = f"Drawing shows {int(drawing_qty)} via {source_label} — {abs(int(diff))} more than BOQ ({int(boq_qty)})."
    else:
        note = f"Drawing shows {int(drawing_qty)} via {source_label} — {abs(int(diff))} fewer than BOQ ({int(boq_qty)})."

    if len(boq_data['items']) > 1:
        sub_parts = []
        for item in boq_data['items']:
            short = item['description'][:35]
            sub_parts.append(f"{short}={int(item['qty'])}")
        note += f" BOQ breakdown: {', '.join(sub_parts)}."

    if etype == 'return_diffuser' and diff > 0:
        note += " Note: block detection may double-count *U16/*U17 inserts — verify against drawing legend."
    elif etype == 'vrf' and diff < 0:
        note += " Note: drawing detects unique VRF labels only — BOQ may list individual modules per system."
    elif etype == 'flow_bar' and diff < 0:
        note += " Note: flow bars often lack distinct block markers — text detection may undercount."
    elif etype == 'volume_control_damper' and len(boq_data['items']) > 1:
        note += " Note: BOQ may list different damper sizes separately — verify each size against drawing."

    return note


def _build_unit_mismatch_note(etype, boq_data, drawing_data, units):
    """Build note for items where BOQ units differ from TraceQ entity count."""
    unit_list = ', '.join(sorted(units))
    boq_qty = boq_data['total_qty']
    drawing_qty = drawing_data.get('count', 0)
    source = drawing_data.get('source', 'detection')
    source_label = _format_source_label(source).lower()

    if drawing_qty > 0:
        base = f"BOQ = {boq_qty:,.1f} {unit_list}. TraceQ found {int(drawing_qty)} entities via {source_label}."
    else:
        base = f"BOQ = {boq_qty:,.1f} {unit_list}. Not detected in drawing."

    if any(u.strip('.') in ('sqm', 'sq.m', 'sq m', 'm2', 'sqft') for u in units):
        base += " BOQ measured in area — duct schedule comparison recommended."
    elif any(u.strip('.') in ('mtrs', 'mtr', 'm', 'lm', 'rm') for u in units):
        base += " BOQ measured in length — direct entity comparison not possible."

    return base


# ═══════════════════════════════════════════════════════════════════
# MAIN COMPARE FUNCTION
# ═══════════════════════════════════════════════════════════════════

def compare_boq_vs_drawing(boq_items, drawing_merged):
    """
    Compare BOQ line items against drawing detection results.
    Returns (comparisons, missing_from_boq) with category-based Trace IDs.
    Status labels: MATCH, DISCREPANCY only (no risk levels).
    Items appear in BOQ order (first-seen equipment type order).
    """
    comparisons = []
    trace_counters = {}  # per-category counters for TQ-[CAT]-[NNN]

    def _make_trace_id(etype):
        prefix = TRACE_PREFIX_MAP.get(etype, 'EQUIP')
        trace_counters[prefix] = trace_counters.get(prefix, 0) + 1
        return f"TQ-{prefix}-{trace_counters[prefix]:03d}"

    # Group BOQ items by equipment type, preserving first-seen order
    boq_by_type = {}
    boq_type_order = []  # preserves BOQ order
    for item in boq_items:
        etype = item['equipment_type']
        if etype is None:
            continue
        if etype not in boq_by_type:
            boq_by_type[etype] = {
                'total_qty': 0,
                'items': [],
                'total_cost': 0,
                'rates': [],
                'units': set(),
            }
            boq_type_order.append(etype)
        boq_by_type[etype]['total_qty'] += item['qty']
        boq_by_type[etype]['items'].append(item)
        if item.get('total'):
            boq_by_type[etype]['total_cost'] += item['total']
        if item.get('rate') and item['rate'] > 0:
            boq_by_type[etype]['rates'].append(item['rate'])
        if item.get('unit'):
            boq_by_type[etype]['units'].add(item['unit'].strip().lower())

    # Build comparison for each BOQ equipment type — IN BOQ ORDER
    matched_drawing_types = set()

    for etype in boq_type_order:
        boq_data = boq_by_type[etype]
        drawing_data = drawing_merged.get(etype, {})
        matched_drawing_types.add(etype)

        boq_qty = boq_data['total_qty']
        drawing_qty = drawing_data.get('count', 0)
        source = drawing_data.get('source', '—')
        rates = boq_data['rates']
        avg_rate = sum(rates) / len(rates) if rates else UAE_UNIT_RATES.get(etype, 0)
        units = boq_data['units']

        has_non_countable = any(u.strip('.') not in ('nos', 'no', 'pcs', 'ea', 'each', 'set', 'sets') for u in units)

        diff = drawing_qty - boq_qty
        exposure = abs(diff) * avg_rate if avg_rate and diff != 0 else 0

        # Determine status — ONLY MATCH or DISCREPANCY (no risk levels)
        is_unit_mismatch = has_non_countable
        if diff == 0 and drawing_qty > 0 and not is_unit_mismatch:
            status = 'MATCH'
            note = "Exact match."
        else:
            status = 'DISCREPANCY'
            if is_unit_mismatch:
                note = _build_unit_mismatch_note(etype, boq_data, drawing_data, units)
            elif drawing_qty == 0:
                note = f"BOQ has {int(boq_qty)} but not detected in drawing. Manual verification required."
            else:
                note = _build_discrepancy_note(etype, boq_data, drawing_data, diff)

        boq_breakdown = _format_boq_breakdown(boq_data['items'])
        name = _format_equipment_name(etype)
        source_label = _format_source_label(source)
        trace_id = _make_trace_id(etype)

        # Always show drawing qty
        if drawing_qty > 0:
            if is_unit_mismatch:
                show_drawing_qty = f"{int(drawing_qty)} entities"
            else:
                show_drawing_qty = int(drawing_qty)
            show_diff = f"{int(diff):+d}" if not is_unit_mismatch else '—'
            variance_pct = f"{abs(diff) / max(boq_qty, 1) * 100:.0f}%" if not is_unit_mismatch and boq_qty > 0 else '—'
        else:
            show_drawing_qty = 'Not Detected'
            show_diff = '—'
            variance_pct = '—'

        if status == 'DISCREPANCY' and exposure > 0 and not is_unit_mismatch:
            show_exposure = f"{exposure:,.0f}"
        else:
            show_exposure = '—'
            if status == 'MATCH':
                exposure = 0

        comparisons.append({
            'Trace ID': trace_id,
            'Equipment': name,
            'BOQ Qty': int(boq_qty) if boq_qty == int(boq_qty) else f"{boq_qty:,.1f}",
            'Drawing Qty': show_drawing_qty,
            'Difference': show_diff,
            'Variance %': variance_pct,
            'Unit': ', '.join(sorted(units)) if units else '—',
            'Risk': status,
            'Exposure (AED)': show_exposure,
            'Notes': note,
            'BOQ Breakdown': boq_breakdown,
            'Detection Source': source_label,
            'Status': status,
            '_exposure_num': exposure if not is_unit_mismatch else 0,
            '_boq_qty': boq_qty,
            '_drawing_qty': drawing_qty if drawing_qty > 0 else None,
            '_diff': diff if not is_unit_mismatch else None,
            '_rate': avg_rate,
            '_is_unit_mismatch': is_unit_mismatch,
            '_equipment_type': etype,
        })

    # Missing from BOQ: items in drawing but not in any BOQ type
    missing_from_boq = []
    for etype, data in sorted(drawing_merged.items()):
        if etype not in matched_drawing_types and data.get('count', 0) > 0:
            trace_id = _make_trace_id(etype) if etype in TRACE_PREFIX_MAP else f"TQ-MISS-{trace_counters.get('MISS', 0) + 1:03d}"
            if etype not in TRACE_PREFIX_MAP:
                trace_counters['MISS'] = trace_counters.get('MISS', 0) + 1
            name = _format_equipment_name(etype)
            source = data.get('source', 'unknown')
            source_label = _format_source_label(source)
            confidence = data.get('confidence', 0)
            qty = data['count']
            unit_rate = UAE_UNIT_RATES.get(etype, 0)
            est_exposure = qty * unit_rate

            missing_from_boq.append({
                'Trace ID': trace_id,
                'Equipment': name,
                'Drawing Qty': qty,
                'Detection': source_label,
                'Detection Source': source_label,
                'Confidence': f"{int(confidence * 100)}%",
                'Notes': f"Found in drawing via {source_label.lower()} but no matching BOQ line item.",
                'Status': 'MISSING FROM BOQ',
                '_equipment_type': etype,
                '_unit_rate': unit_rate,
                '_est_exposure': est_exposure,
            })

    return comparisons, missing_from_boq


# ═══════════════════════════════════════════════════════════════════
# CONVENIENCE: Run full comparison from file paths
# ═══════════════════════════════════════════════════════════════════

def run_comparison(dxf_path, boq_path):
    """
    Run the full TraceQ comparison pipeline from file paths.
    Returns (comparisons, missing_from_boq, engine_result)
    """
    import sys
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from traceq_engine import TraceQEngine

    # Parse BOQ
    boq_items = parse_boq_file(boq_path)
    print(f"[TraceQ Compare] Parsed {len(boq_items)} BOQ items")

    # Run engine
    engine = TraceQEngine()
    result = engine.analyze(dxf_path)

    # Compare
    comparisons, missing = compare_boq_vs_drawing(boq_items, result.merged)
    print(f"[TraceQ Compare] {len(comparisons)} comparisons, {len(missing)} missing from BOQ")

    return comparisons, missing, result


if __name__ == '__main__':
    """Quick test: run S5 comparison and print results."""
    import sys
    if len(sys.argv) < 3:
        print("Usage: python traceq_compare.py <drawing.dxf> <boq.xlsx>")
        sys.exit(1)

    comparisons, missing, result = run_comparison(sys.argv[1], sys.argv[2])

    print(f"\n{'='*80}")
    print(f"COMPARISON RESULTS")
    print(f"{'='*80}")
    for c in comparisons:
        print(f"  {c['Trace ID']} | {c['Equipment']:25s} | BOQ: {c['BOQ Qty']:>8} | Draw: {str(c['Drawing Qty']):>12} | {c['Status']}")
    print(f"\nMISSING FROM BOQ:")
    for m in missing:
        print(f"  {m['Trace ID']} | {m['Equipment']:25s} | Draw: {m['Drawing Qty']:>5} | {m['Status']}")
